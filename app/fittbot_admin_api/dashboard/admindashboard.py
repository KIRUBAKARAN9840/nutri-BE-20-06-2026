# Backend Implementation Guide for Dashboard APIs
from fastapi import APIRouter, Depends, HTTPException, FastAPI, Query
from pydantic import BaseModel
from typing import Optional
from sqlalchemy.ext.asyncio import AsyncSession
from datetime import datetime, timedelta
from sqlalchemy import func, and_, select, distinct, or_, text, String, desc, cast, Integer
from sqlalchemy.sql.expression import literal

from sqlalchemy.orm import Session
import io
import pandas as pd
from fastapi.responses import StreamingResponse
from openpyxl.styles import PatternFill, Font, Alignment

# Import centralized revenue service for consistency
from app.fittbot_admin_api.revenue_service import (
    get_revenue_breakdown,
    get_detailed_revenue_with_breakdowns,
    paise_to_rupees,
    get_paying_users_count
)

from app.models.fittbot_models import (
    Client, Gym, ClientToken, OwnerToken, GymOwner, RewardInterest, RewardProgramOptIn,
    SessionSetting, GymPlans, GymStudiosPic, GymOnboardingPics,
    SessionBookingDay, SessionBooking, ClassSession, ClientFittbotAccess, ActiveUser, SessionPurchase,
    GymMateProfile
)
from app.models.adminmodels import(TicketAssignment, Employees, Admins)
from app.models.async_database import get_async_db
from app.fittbot_admin_api.auth.authentication import get_current_admin_from_cookie
from app.fittbot_api.v1.payments.models.subscriptions import Subscription
from app.fittbot_api.v1.payments.models.catalog import CatalogProduct
from app.fittbot_api.v1.payments.models.payments import Payment
from app.fittbot_api.v1.payments.models.orders import Order, OrderItem
from app.models.nutrition_models import NutritionEligibility
from app.models.dailypass_models import DailyPassPricing, get_dailypass_session, DailyPass

router = APIRouter(prefix="/api/admin/dashboard", tags=["AdminDashboard"])

async def get_monthly_active_users(
    db: AsyncSession, 
    start_date: Optional[datetime.date] = None, 
    end_date: Optional[datetime.date] = None,
    is_overall: bool = False
) -> int:
    """
    Get count of distinct client_ids from active_users table
    where created_at is within the specified date range.
    Requires only 1+ login (same logic as /users-stats Active Users).
    Excludes users from gym_id = 1.
    """
    try:
        active_filters = []
        if not is_overall:
            if start_date:
                active_filters.append(ActiveUser.created_at >= datetime.combine(start_date, datetime.min.time()))
            else:
                # Default to last 30 days
                thirty_days_ago = datetime.now().date() - timedelta(days=30)
                active_filters.append(ActiveUser.created_at >= datetime.combine(thirty_days_ago, datetime.min.time()))
        
        if end_date:
            active_filters.append(ActiveUser.created_at <= datetime.combine(end_date, datetime.max.time()))

        active_subquery = select(ActiveUser.client_id).join(
            Client, ActiveUser.client_id == Client.client_id
        ).where(
            and_(
                *active_filters,
                or_(Client.gym_id != 1, Client.gym_id.is_(None))
            )
        )

        stmt = select(
            func.coalesce(func.count(func.distinct(ActiveUser.client_id)), 0)
        ).where(
            ActiveUser.client_id.in_(active_subquery)
        )

        result = await db.execute(stmt)
        count = result.scalar()
        return int(count) if count is not None else 0
    except Exception as e:
        print(f"[MONTHLY_ACTIVE_USERS] Error fetching active users: {e}")
        return 0

async def get_total_paying_users(db: AsyncSession) -> int:
    """
    Get count of distinct customer_id from payments table using common helper.
    """
    return await get_paying_users_count(db)

async def get_fittbot_metrics(db: AsyncSession, filter_type='month'):

    today = datetime.now().date()

    # Total users based on filter
    stmt = select(func.count()).select_from(Client).filter(
        func.date(Client.created_at) == today
    )
    result = await db.execute(stmt)
    total_users_today = result.scalar() or 0

    # Yesterday users
    stmt = select(func.count()).select_from(Client).filter(
        func.date(Client.created_at) == today - timedelta(days=1)
    )
    result = await db.execute(stmt)
    total_users_yesterday = result.scalar() or 0

    stmt = select(func.count()).select_from(Client).filter(
        Client.created_at >= today - timedelta(days=7),
        Client.created_at < today + timedelta(days=1)
    )
    result = await db.execute(stmt)
    total_users_week = result.scalar() or 0

    stmt = select(func.count()).select_from(Client).filter(
        Client.created_at >= today - timedelta(days=30),
        Client.created_at < today + timedelta(days=1)
    )
    result = await db.execute(stmt)
    total_users_month = result.scalar() or 0

    stmt = select(func.count()).select_from(Client)
    result = await db.execute(stmt)
    total_users_overall = result.scalar() or 0

    # Retrieve first user registration date
    stmt_min = select(func.min(Client.created_at))
    result_min = await db.execute(stmt_min)
    first_user_date = result_min.scalar()
    
    first_user_date_str = None
    if first_user_date:
        if isinstance(first_user_date, datetime):
            first_user_date_str = first_user_date.date().isoformat()
        elif isinstance(first_user_date, date):
            first_user_date_str = first_user_date.isoformat()
        else:
            first_user_date_str = str(first_user_date)

    # Revenue calculation
    revenue_data = await calculate_revenue(db, today)

    # Subscribed users calculation
    subscribed_users_data = await calculate_subscribed_users(db, today)

    # Monthly revenue trends for last 6 months
    monthly_revenue_trends = await calculate_monthly_revenue_trends(db, today)

    # Monthly active users calculations for all filters
    active_users_today = await get_monthly_active_users(db, today, today)
    active_users_yesterday = await get_monthly_active_users(db, today - timedelta(days=1), today - timedelta(days=1))
    active_users_week = await get_monthly_active_users(db, today - timedelta(days=7), today)
    active_users_month = await get_monthly_active_users(db, today - timedelta(days=30), today)
    
    first_day_of_last_month = (today.replace(day=1) - timedelta(days=1)).replace(day=1)
    last_day_of_last_month = today.replace(day=1) - timedelta(days=1)
    active_users_last_month = await get_monthly_active_users(db, first_day_of_last_month, last_day_of_last_month)
    
    first_day_of_current_month = today.replace(day=1)
    active_users_current_month = await get_monthly_active_users(db, first_day_of_current_month, today)
    
    active_users_overall = await get_monthly_active_users(db, is_overall=True)

    # Total paying users (distinct customer_id from payments table)
    total_paying_users = await get_total_paying_users(db)

    return {
        "totalUsers": {
            "today": total_users_today,
            "yesterday": total_users_yesterday,
            "week": total_users_week,
            "month": total_users_month,
            "overall": total_users_overall,
            "first_user_date": first_user_date_str
        },
        "revenue": revenue_data,
        "subscribedUsers": subscribed_users_data,
        "monthlyActiveUsers": {
            "today": active_users_today,
            "yesterday": active_users_yesterday,
            "week": active_users_week,
            "month": active_users_month,
            "lastMonth": active_users_last_month,
            "currentMonth": active_users_current_month,
            "overall": active_users_overall
        },
        "totalPayingUsers": total_paying_users,
        "monthlyRevenueTrends": monthly_revenue_trends
    }

async def get_fittbot_metrics_custom(db: AsyncSession, start_date: str, end_date: str):
    """
    Get Fittbot metrics for custom date range
    """
    try:
        start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    if start_date_obj > end_date_obj:
        raise HTTPException(status_code=400, detail="Start date must be before end date")

    # Total users in custom range
    stmt = select(func.count()).select_from(Client).filter(
        func.date(Client.created_at) >= start_date_obj,
        func.date(Client.created_at) <= end_date_obj
    )
    result = await db.execute(stmt)
    total_users_custom = result.scalar() or 0

    # Revenue in custom range - using centralized revenue service
    revenue_data = await get_revenue_breakdown(
        db=db,
        start_date=start_date_obj,
        end_date=end_date_obj,
        exclude_gym_id_one=True
    )
    total_revenue = revenue_data.total_revenue

    # Subscribed users in custom range (subscriptions that started in custom range and are still active)
    now = datetime.now()
    stmt = select(func.count(distinct(Subscription.customer_id))).filter(
        Subscription.provider.in_(['razorpay_pg', 'google_play']),
        Subscription.status != 'pending',
        func.date(Subscription.active_from) >= start_date_obj,
        func.date(Subscription.active_from) <= end_date_obj,
        Subscription.active_until >= now
    )
    result = await db.execute(stmt)
    subscribed_custom = result.scalar() or 0

    # Monthly active users - use start_date and end_date to get active users in custom range
    monthly_active_users_custom = await get_monthly_active_users(db, start_date_obj, end_date_obj)

    # Total paying users (distinct customer_id from payments table)
    total_paying_users_custom = await get_total_paying_users(db)

    return {
        "totalUsers": {
            "today": 0,
            "week": 0,
            "month": 0,
            "overall": 0,
            "custom": total_users_custom
        },
        "revenue": {
            "today": "₹0",
            "week": "₹0",
            "month": "₹0",
            "overall": "₹0",
            "custom": f"₹{paise_to_rupees(total_revenue):,.0f}"
        },
        "subscribedUsers": {
            "today": 0,
            "week": 0,
            "month": 0,
            "overall": 0,
            "custom": subscribed_custom
        },
        "monthlyActiveUsers": {
            "today": 0,
            "yesterday": 0,
            "week": 0,
            "month": 0,
            "lastMonth": 0,
            "currentMonth": 0,
            "overall": 0,
            "custom": monthly_active_users_custom
        },
        "totalPayingUsers": total_paying_users_custom,
        "monthlyRevenueTrends": []
    }

async def get_business_metrics_custom(db: AsyncSession, start_date: str, end_date: str):
    """
    Get Business metrics for custom date range
    """
    try:
        start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    if start_date_obj > end_date_obj:
        raise HTTPException(status_code=400, detail="Start date must be before end date")

    # Gym Owners in custom range
    stmt = select(func.count()).select_from(GymOwner).filter(
        func.date(GymOwner.created_at) >= start_date_obj,
        func.date(GymOwner.created_at) <= end_date_obj
    )
    result = await db.execute(stmt)
    gym_owners_custom = result.scalar() or 0

    # Gyms in custom range
    stmt = select(func.count()).select_from(Gym).filter(
        func.date(Gym.created_at) >= start_date_obj,
        func.date(Gym.created_at) <= end_date_obj
    )
    result = await db.execute(stmt)
    gyms_custom = result.scalar() or 0

    # Daily pass enabled gyms (total, not affected by date range)
    stmt = select(func.count()).select_from(Gym).filter(
        Gym.dailypass == True
    )
    result = await db.execute(stmt)
    daily_pass_gyms = result.scalar() or 0

    # Verified gyms (type contains "green") - total, not affected by date range
    # Count total gyms (excluding "dummy" type)
    stmt = select(func.count()).select_from(Gym).filter(
        or_(
            Gym.type.notlike("%dummy%"),
            Gym.type.is_(None)
        )
    )
    result = await db.execute(stmt)
    total_gyms_count = result.scalar() or 0

    # Count gyms where type contains "green"
    stmt = select(func.count()).select_from(Gym).filter(
        Gym.type.like("%green%")
    )
    result = await db.execute(stmt)
    verified_gyms_count = result.scalar() or 0

    # Unverified gyms (type contains "hold" or "red")
    stmt = select(func.count()).select_from(Gym).filter(
        or_(
            Gym.type.like("%hold%"),
            Gym.type.like("%red%")
        )
    )
    result = await db.execute(stmt)
    unverified_gyms_count = result.scalar() or 0

    # Count gyms where type = 'red'
    stmt = select(func.count()).select_from(Gym).filter(
        Gym.type == 'red'
    )
    result = await db.execute(stmt)
    red_gyms_count = result.scalar() or 0

    # Count gyms where type = 'hold'
    stmt = select(func.count()).select_from(Gym).filter(
        Gym.type == 'hold'
    )
    result = await db.execute(stmt)
    hold_gyms_count = result.scalar() or 0

    return {
        "gymOwners": {
            "today": 0,
            "week": 0,
            "month": 0,
            "overall": 0,
            "custom": gym_owners_custom
        },
        "gyms": {
            "today": 0,
            "week": 0,
            "month": 0,
            "overall": 0,
            "custom": gyms_custom
        },
        "dailyPassGyms": daily_pass_gyms,
        "verifiedGyms": {
            "verified": verified_gyms_count,
            "total": total_gyms_count
        },
        "unverifiedGyms": unverified_gyms_count,
        "unverifiedSplitup": {
            "red": red_gyms_count,
            "hold": hold_gyms_count
        }
    }

async def calculate_revenue(db: AsyncSession, today):
    """
    Calculate revenue for different time periods using centralized revenue service.

    Returns formatted revenue strings in rupees.
    """
    # Calculate revenue for different time periods using centralized service
    revenue_today = await get_revenue_breakdown(
        db=db,
        start_date=today,
        end_date=today,
        exclude_gym_id_one=True
    )

    revenue_week = await get_revenue_breakdown(
        db=db,
        start_date=today - timedelta(days=7),
        end_date=today,
        exclude_gym_id_one=True
    )

    revenue_month = await get_revenue_breakdown(
        db=db,
        start_date=today - timedelta(days=30),
        end_date=today,
        exclude_gym_id_one=True
    )

    # Overall revenue (all time)
    revenue_overall = await get_revenue_breakdown(
        db=db,
        start_date=datetime(2020, 1, 1).date(),
        end_date=today,
        exclude_gym_id_one=True
    )

    # Convert from paise to rupees and format
    return {
        "today": f"₹{paise_to_rupees(revenue_today.total_revenue):,.0f}",
        "week": f"₹{paise_to_rupees(revenue_week.total_revenue):,.0f}",
        "month": f"₹{paise_to_rupees(revenue_month.total_revenue):,.0f}",
        "overall": f"₹{paise_to_rupees(revenue_overall.total_revenue):,.0f}"
    }

async def calculate_subscribed_users(db: AsyncSession, today):
 
    now = datetime.now()

    # Today's subscribed users (subscriptions that started today and are still active)
    stmt = select(func.count(distinct(Subscription.customer_id))).filter(
        Subscription.provider.in_(['razorpay_pg', 'google_play']),
        Subscription.status != 'pending',
        func.date(Subscription.active_from) == today,
        Subscription.active_until >= now
    )
    result = await db.execute(stmt)
    subscribed_today = result.scalar() or 0

    # This week's subscribed users (subscriptions that started this week and are still active)
    stmt = select(func.count(distinct(Subscription.customer_id))).filter(
        Subscription.provider.in_(['razorpay_pg', 'google_play']),
        Subscription.status != 'pending',
        Subscription.active_from >= today - timedelta(days=7),
        Subscription.active_from < today + timedelta(days=1),
        Subscription.active_until >= now
    )
    result = await db.execute(stmt)
    subscribed_week = result.scalar() or 0

    # This month's subscribed users (subscriptions that started this month and are still active)
    stmt = select(func.count(distinct(Subscription.customer_id))).filter(
        Subscription.provider.in_(['razorpay_pg', 'google_play']),
        Subscription.status != 'pending',
        Subscription.active_from >= today - timedelta(days=30),
        Subscription.active_from < today + timedelta(days=1),
        Subscription.active_until >= now
    )
    result = await db.execute(stmt)
    subscribed_month = result.scalar() or 0

    # Overall subscribed users (all active subscriptions)
    stmt = select(func.count(distinct(Subscription.customer_id))).filter(
        Subscription.provider.in_(['razorpay_pg', 'google_play']),
        Subscription.status != 'pending',
        Subscription.active_until >= now
    )
    result = await db.execute(stmt)
    subscribed_overall = result.scalar() or 0

    return {
        "today": subscribed_today,
        "week": subscribed_week,
        "month": subscribed_month,
        "overall": subscribed_overall
    }

async def calculate_monthly_revenue_trends(db: AsyncSession, today):
   
    import calendar

    monthly_data = []
    current_date = datetime.now()

    # Get last 6 months
    for i in range(5, -1, -1):
        # Calculate the month
        target_month = current_date.month - i
        target_year = current_date.year

        # Handle year rollover
        while target_month <= 0:
            target_month += 12
            target_year -= 1

        # Get first and last day of the month
        first_day = datetime(target_year, target_month, 1).date()
        last_day = datetime(target_year, target_month, calendar.monthrange(target_year, target_month)[1]).date()

        # Calculate revenue for this month
        stmt = select(
            func.coalesce(func.sum(CatalogProduct.base_amount_minor), 0)
        ).select_from(Subscription).join(
            CatalogProduct, Subscription.product_id == CatalogProduct.sku
        ).filter(
            Subscription.provider.in_(['razorpay_pg', 'google_play']),
            Subscription.status != 'pending',
            Subscription.active_from >= first_day,
            Subscription.active_from <= last_day
        )
        result = await db.execute(stmt)
        monthly_revenue = result.scalar() or 0

        # Convert from paise to thousands (1000 rupees = 100000 paise)
        revenue_in_thousands = monthly_revenue / 100000

        # Get month abbreviation
        month_abbr = calendar.month_abbr[target_month]

        monthly_data.append({
            "month": month_abbr,
            "revenue": round(revenue_in_thousands, 2)
        })

    return monthly_data

async def get_plans_metrics(db: AsyncSession):
    """
    Get Plans metrics (nutritionist plan purchases data)

    NEW LOGIC: Counts payments from payments.payments table
    where payment_metadata['flow'] = 'nutrition_purchase_googleplay' or 'nutrition_package_razorpay'
    """
    # EXCLUDED_CONTACTS for test/internal accounts
    EXCLUDED_CONTACTS = ["7373675762", "9486987082", "8667458723", "9840633149", "8667427956", "8667488723", "7975847236"]

    # Nutritionist Plans - count payments where flow = 'nutrition_purchase_googleplay' or 'nutrition_package_razorpay'
    # Exclude internal/test contacts to match the inner page logic
    stmt = (
        select(func.count(distinct(Payment.customer_id)))
        .select_from(Payment)
        .outerjoin(Client, Payment.customer_id == Client.client_id)
        .join(
            NutritionEligibility,
            cast(Payment.order_id, String) == NutritionEligibility.source_id
        )
        .where(
            NutritionEligibility.source_type == "fymble_purchase",
            Payment.status == "captured",
            or_(
                func.json_extract(Payment.payment_metadata, '$.flow') == 'nutrition_purchase_googleplay',
                func.json_extract(Payment.payment_metadata, '$.flow') == 'nutrition_package_razorpay',
                func.json_extract(Payment.payment_metadata, '$.flow') == 'basic_nutrition_plan',
                func.json_extract(Payment.payment_metadata, '$.flow') == 'expert_nutrition_plan',
                func.json_extract(Payment.payment_metadata, '$.flow') == 'elite_nutrition_plan'
            ),
            ~Client.contact.in_(EXCLUDED_CONTACTS)
        )
    )
    result = await db.execute(stmt)
    nutritionist_total = result.scalar() or 0

    # Complimentary Plans - count from NutritionEligibility where source_type != 'fymble_purchase'
    stmt_comp = (
        select(func.count(distinct(NutritionEligibility.client_id)))
        .select_from(NutritionEligibility)
        .outerjoin(Client, NutritionEligibility.client_id == Client.client_id)
        .where(
            NutritionEligibility.source_type != "fymble_purchase",
            ~Client.contact.in_(EXCLUDED_CONTACTS)
        )
    )
    result_comp = await db.execute(stmt_comp)
    complimentary_total = result_comp.scalar() or 0

    # For one-time purchases, all are counted as "total"
    # Gold/Platinum/Diamond breakdown is not applicable for one-time purchases
    return {
        "complimentary": complimentary_total,
        "fittbotSubscriptions": {
            "total": nutritionist_total,
            "gold": nutritionist_total,  # All plans count toward each category
            "platinum": nutritionist_total,
            "diamond": nutritionist_total
        }
    }

async def get_business_metrics(db: AsyncSession, filter_type='month'):
    """
    Get Business metrics (gym-related data)
    """
    today = datetime.now().date()

    # Gym Owners count based on filter
    stmt = select(func.count()).select_from(GymOwner).filter(
        func.date(GymOwner.created_at) == today
    )
    result = await db.execute(stmt)
    gym_owners_today = result.scalar() or 0

    stmt = select(func.count()).select_from(GymOwner).filter(
        GymOwner.created_at >= today - timedelta(days=7),
        GymOwner.created_at < today + timedelta(days=1)
    )
    result = await db.execute(stmt)
    gym_owners_week = result.scalar() or 0

    stmt = select(func.count()).select_from(GymOwner).filter(
        GymOwner.created_at >= today - timedelta(days=30),
        GymOwner.created_at < today + timedelta(days=1)
    )
    result = await db.execute(stmt)
    gym_owners_month = result.scalar() or 0

    stmt = select(func.count()).select_from(GymOwner)
    result = await db.execute(stmt)
    gym_owners_overall = result.scalar() or 0

    # Gyms count based on filter
    stmt = select(func.count()).select_from(Gym).filter(
        func.date(Gym.created_at) == today
    )
    result = await db.execute(stmt)
    gyms_today = result.scalar() or 0

    stmt = select(func.count()).select_from(Gym).filter(
        Gym.created_at >= today - timedelta(days=7),
        Gym.created_at < today + timedelta(days=1)
    )
    result = await db.execute(stmt)
    gyms_week = result.scalar() or 0

    stmt = select(func.count()).select_from(Gym).filter(
        Gym.created_at >= today - timedelta(days=30),
        Gym.created_at < today + timedelta(days=1)
    )
    result = await db.execute(stmt)
    gyms_month = result.scalar() or 0

    stmt = select(func.count()).select_from(Gym)
    result = await db.execute(stmt)
    gyms_overall = result.scalar() or 0

    # Daily pass enabled gyms (no time filter, just total count where dailypass = 1/True)
    stmt = select(func.count()).select_from(Gym).filter(
        Gym.dailypass == True
    )
    result = await db.execute(stmt)
    daily_pass_gyms = result.scalar() or 0

    # Verified gyms (type contains "green")
    # Count total gyms (excluding "dummy" type)
    stmt = select(func.count()).select_from(Gym).filter(
        or_(
            Gym.type.notlike("%dummy%"),
            Gym.type.is_(None)
        )
    )
    result = await db.execute(stmt)
    total_gyms_count = result.scalar() or 0

    # Count gyms where type contains "green"
    stmt = select(func.count()).select_from(Gym).filter(
        Gym.type.like("%green%")
    )
    result = await db.execute(stmt)
    verified_gyms_count = result.scalar() or 0

    # Unverified gyms (type contains "hold" or "red")
    stmt = select(func.count()).select_from(Gym).filter(
        or_(
            Gym.type.like("%hold%"),
            Gym.type.like("%red%")
        )
    )
    result = await db.execute(stmt)
    unverified_gyms_count = result.scalar() or 0

    # Count gyms where type = 'red'
    stmt = select(func.count()).select_from(Gym).filter(
        Gym.type == 'red'
    )
    result = await db.execute(stmt)
    red_gyms_count = result.scalar() or 0

    # Count gyms where type = 'hold'
    stmt = select(func.count()).select_from(Gym).filter(
        Gym.type == 'hold'
    )
    result = await db.execute(stmt)
    hold_gyms_count = result.scalar() or 0

    return {
        "gymOwners": {
            "today": gym_owners_today,
            "week": gym_owners_week,
            "month": gym_owners_month,
            "overall": gym_owners_overall
        },
        "gyms": {
            "today": gyms_today,
            "week": gyms_week,
            "month": gyms_month,
            "overall": gyms_overall
        },
        "dailyPassGyms": daily_pass_gyms,
        "verifiedGyms": {
            "verified": verified_gyms_count,
            "total": total_gyms_count
        },
        "unverifiedGyms": unverified_gyms_count,
        "unverifiedSplitup": {
            "red": red_gyms_count,
            "hold": hold_gyms_count
        }
    }

async def get_support_tickets(db: AsyncSession):
    """
    Get Support tickets data
    """
    today = datetime.now().date()

    stmt = select(func.count()).select_from(ClientToken)
    result = await db.execute(stmt)
    total_client_tickets = result.scalar() or 0

    stmt = select(func.count()).select_from(OwnerToken)
    result = await db.execute(stmt)
    total_gym_tickets = result.scalar() or 0

    stmt = select(func.count()).select_from(ClientToken).filter(
        ClientToken.resolved == False
    )
    result = await db.execute(stmt)
    unresolved_client_tickets = result.scalar() or 0

    stmt = select(func.count()).select_from(OwnerToken).filter(
        OwnerToken.resolved == False
    )
    result = await db.execute(stmt)
    unresolved_gym_tickets = result.scalar() or 0

    # Client tokens resolved today
    stmt_client = select(func.count()).select_from(ClientToken).filter(
        func.date(ClientToken.updated_at) == today,
        ClientToken.resolved == True
    )
    result_client = await db.execute(stmt_client)
    resolved_client_today = result_client.scalar() or 0

    # Owner tokens resolved today
    stmt_owner = select(func.count()).select_from(OwnerToken).filter(
        func.date(OwnerToken.updated_at) == today,
        OwnerToken.resolved == True
    )
    result_owner = await db.execute(stmt_owner)
    resolved_owner_today = result_owner.scalar() or 0

    resolved_today = resolved_client_today + resolved_owner_today

    return {
        "totalTickets": {
            "gym": total_gym_tickets,
            "client": total_client_tickets
        },
        "unresolvedTickets": {
            "gym": unresolved_gym_tickets,
            "client": unresolved_client_tickets
        },
        "resolvedToday": resolved_today
    }

async def get_rewards_metrics(db: AsyncSession):
    """
    Get Rewards metrics from reward_interest table
    """
    # Total count of reward_interest records
    stmt = select(func.count()).select_from(RewardInterest)
    result = await db.execute(stmt)
    total_count = result.scalar() or 0

    # Count of interested (where interested = True)
    stmt = select(func.count()).select_from(RewardInterest).filter(
        RewardInterest.interested == True
    )
    result = await db.execute(stmt)
    interested_count = result.scalar() or 0

    return {
        "total": total_count,
        "interested": interested_count
    }

async def get_reward_program_participants(db: AsyncSession):
    """
    Get Reward Program Participants count from reward_program_opt_ins table
    Counts the total number of records based on client_id
    """
    # Count distinct client_ids from reward_program_opt_ins
    stmt = select(func.count(distinct(RewardProgramOptIn.client_id)))
    result = await db.execute(stmt)
    total_participants = result.scalar() or 0

    return {
        "totalParticipants": total_participants
    }

async def get_gym_plans_metrics(db: AsyncSession):
    """
    Get Gym Plans metrics - count of gyms with session plans, membership plans, and daily pass pricing
    """
    # Get total gym count first
    stmt = select(func.count()).select_from(Gym)
    result = await db.execute(stmt)
    total_gyms = result.scalar() or 0

    # Count gyms with session plans
    stmt = select(func.count(distinct(SessionSetting.gym_id)))
    result = await db.execute(stmt)
    gyms_with_session_plans = result.scalar() or 0

    # Count gyms with membership plans
    stmt = select(func.count(distinct(GymPlans.gym_id)))
    result = await db.execute(stmt)
    gyms_with_membership_plans = result.scalar() or 0

    stmt = select(Gym.gym_id)
    result = await db.execute(stmt)
    all_gym_ids = [row[0] for row in result.all()]
    all_gym_id_strings = [str(gym_id) for gym_id in all_gym_ids]

    # Get all gym_id strings from DailyPassPricing that match our gyms
    stmt = select(DailyPassPricing.gym_id).filter(
        DailyPassPricing.gym_id.in_(all_gym_id_strings)
    )
    result = await db.execute(stmt)
    daily_pass_gym_ids = result.all()

    # Count unique gym IDs (convert to int to avoid duplicates from string representation)
    gyms_with_daily_pass = len(set([int(gym_id[0]) for gym_id in daily_pass_gym_ids]))

    return {
        "sessionPlans": gyms_with_session_plans,
        "membershipPlans": gyms_with_membership_plans,
        "dailyPass": gyms_with_daily_pass,
        "totalGyms": total_gyms
    }

async def get_gym_photos_metrics(db: AsyncSession):
    """
    Calculate gym photos metrics using mutually exclusive logic.
    Logic matches the gymdetails page:
    1. Verified Studio: Gyms with photos in gym_studios_pic (priority)
    2. Pending Photo verification: Gyms with photos ONLY in gym_onboarding_pics (NO studio photos)
    3. Photos Not Uploaded: Gyms with NEITHER studio NOR onboarding photos
    """

    # Get all gym IDs
    stmt = select(Gym.gym_id)
    result = await db.execute(stmt)
    all_gym_ids = [row[0] for row in result.all()]

    # Get gym IDs with studio photos
    stmt = select(GymStudiosPic.gym_id)
    result = await db.execute(stmt)
    gym_ids_with_studio = set([row[0] for row in result.all()])

    # Get gym IDs with onboarding photos
    stmt = select(GymOnboardingPics.gym_id)
    result = await db.execute(stmt)
    gym_ids_with_onboarding = set([row[0] for row in result.all()])

    # Verified Studio: Gyms with studio photos (priority given to studio)
    verified_studio_count = len(gym_ids_with_studio)

    # Pending Photo verification: Gyms with ONLY onboarding photos (excluding those with studio photos)
    gym_ids_only_onboarding = gym_ids_with_onboarding - gym_ids_with_studio
    pending_verification_count = len(gym_ids_only_onboarding)

    # Photos Not Uploaded: Gyms with neither studio nor onboarding photos
    gym_ids_with_any_photos = gym_ids_with_studio.union(gym_ids_with_onboarding)
    no_uploads_count = len(all_gym_ids) - len(gym_ids_with_any_photos)

    return {
        "studio": verified_studio_count,
        "onboard": pending_verification_count,
        "noUploads": no_uploads_count
    }

async def get_gymmate_metrics(db: AsyncSession):
    """
    Get GYM Mate metrics - count of profiles where onboarding_completed = 1/True
    and active user averages for GymMate members (Monthly, Weekly, Daily).
    """
    from calendar import monthrange
    from datetime import datetime, date, timedelta, timezone
    
    # 1. Total GymMate profiles completed onboarding
    stmt = select(func.count()).select_from(GymMateProfile).filter(
        GymMateProfile.onboarding_completed == True
    )
    result = await db.execute(stmt)
    total_count = result.scalar() or 0
    
    # 2. Monthly, Weekly, Daily active user counts for GymMate members
    today_utc = datetime.now(timezone.utc).date()
    
    async def get_avg_gymmate_active_count(s_date, e_date):
        try:
            gymmate_sub = select(GymMateProfile.client_id).where(
                GymMateProfile.onboarding_completed == True
            )
            
            sub = select(ActiveUser.client_id).join(
                Client, ActiveUser.client_id == Client.client_id
            ).where(
                and_(
                    func.date(ActiveUser.created_at) >= s_date,
                    func.date(ActiveUser.created_at) <= e_date,
                    or_(Client.gym_id != 1, Client.gym_id.is_(None)),
                    ActiveUser.client_id.in_(gymmate_sub)
                )
            )
            q = select(
                func.coalesce(func.count(func.distinct(ActiveUser.client_id)), 0)
            ).where(ActiveUser.client_id.in_(sub))
            r = await db.execute(q)
            return int(r.scalar() or 0)
        except Exception as e:
            print(f"[DASHBOARD] Error in get_avg_gymmate_active_count: {e}")
            return 0

    # MAU — last 3 completed calendar months
    monthly_counts = []
    current_year = today_utc.year
    current_month = today_utc.month
    for i in range(3):
        month_index = current_month - 1 - i
        yr = current_year
        while month_index < 0:
            month_index += 12
            yr -= 1
        m_start = date(yr, month_index + 1, 1)
        _, last_day = monthrange(yr, month_index + 1)
        m_end = date(yr, month_index + 1, last_day)
        monthly_counts.append(await get_avg_gymmate_active_count(m_start, m_end))
    monthly_average = sum(monthly_counts) / 3

    # WAU — last 3 completed 7-day windows
    weekly_counts = []
    yesterday_utc = today_utc - timedelta(days=1)
    for i in range(3):
        w_end = yesterday_utc - timedelta(weeks=i)
        w_start = w_end - timedelta(days=6)
        weekly_counts.append(await get_avg_gymmate_active_count(w_start, w_end))
    weekly_average = sum(weekly_counts) / 3

    # DAU — last 3 fully completed days
    daily_counts = []
    for i in range(3):
        day = today_utc - timedelta(days=i + 1)
        daily_counts.append(await get_avg_gymmate_active_count(day, day))
    daily_average = sum(daily_counts) / 3
    
    return {
        "total": total_count,
        "monthly_active_users": round(monthly_average, 0),
        "weekly_active_users": round(weekly_average, 0),
        "daily_active_users": round(daily_average, 0)
    }

@router.get("/overview")
async def get_dashboard_overview(
    fittbot_filter: str = "month",
    business_filter: str = "month",
    custom_start_date: str = None,
    custom_end_date: str = None,
    db: AsyncSession = Depends(get_async_db)
):
    try:
        # Use custom date range if provided for fittbot
        if custom_start_date and custom_end_date and fittbot_filter == "custom":
            fittbot_data = await get_fittbot_metrics_custom(db, custom_start_date, custom_end_date)
        else:
            fittbot_data = await get_fittbot_metrics(db, fittbot_filter)

        # Use custom date range if provided for business
        if custom_start_date and custom_end_date and business_filter == "custom":
            business_data = await get_business_metrics_custom(db, custom_start_date, custom_end_date)
        else:
            business_data = await get_business_metrics(db, business_filter)

        plans_data = await get_plans_metrics(db)
        support_data = await get_support_tickets(db)
        rewards_data = await get_rewards_metrics(db)
        gym_plans_data = await get_gym_plans_metrics(db)
        gym_photos_data = await get_gym_photos_metrics(db)
        reward_program_data = await get_reward_program_participants(db)
        gymmate_data = await get_gymmate_metrics(db)

        return {
            "success": True,
            "data": {
                "fittbot": fittbot_data,
                "business": business_data,
                "plans": plans_data,
                "support": support_data,
                "rewards": rewards_data,
                "gymPlans": gym_plans_data,
                "gymPhotos": gym_photos_data,
                "rewardProgram": reward_program_data,
                "gymMate": gymmate_data
            },
            "message": "Dashboard data fetched successfully"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/fittbot-metrics")
async def get_fittbot_metrics_endpoint(filter: str = "month", db: AsyncSession = Depends(get_async_db)):
    try:
        data = await get_fittbot_metrics(db, filter)
        return {
            "success": True,
            "data": data,
            "message": "Fittbot metrics fetched successfully"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/business-metrics")
async def get_business_metrics_endpoint(filter: str = "month", db: AsyncSession = Depends(get_async_db)):
    try:
        data = await get_business_metrics(db, filter)
        return {
            "success": True,
            "data": data,
            "message": "Business metrics fetched successfully"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/support-tickets")
async def get_support_tickets_endpoint(db: AsyncSession = Depends(get_async_db)):
    try:
        data = await get_support_tickets(db)
        return {
            "success": True,
            "data": data,
            "message": "Support tickets data fetched successfully"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/revenue-analytics")
async def get_revenue_analytics(
    start_date: str = None,
    end_date: str = None,
    source: str = None,
    gym_id: int = None,
    db: AsyncSession = Depends(get_async_db)
):
    """
    Get revenue analytics data using centralized revenue service.

    Provides:
    - Total revenue
    - Source-wise breakdown
    - Daily revenue over time
    - Gym-wise revenue breakdown
    """
    try:
        # Parse dates if provided
        if start_date:
            start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
        else:
            start_date_obj = datetime(2020, 1, 1).date()

        if end_date:
            end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
        else:
            end_date_obj = datetime.now().date()

        # Use centralized revenue service for detailed breakdowns
        revenue_data = await get_detailed_revenue_with_breakdowns(
            db=db,
            start_date=start_date_obj,
            end_date=end_date_obj,
            source=source,
            specific_gym_id=gym_id,
            exclude_gym_id_one=True
        )

        # Build response data
        analytics_data = {
            "totalRevenue": revenue_data.total_revenue,
            "sourceSplit": {
                "daily_pass": revenue_data.source_split["daily_pass"],
                "sessions": revenue_data.source_split["sessions"],
                "fittbot_subscription": revenue_data.source_split["fittbot_subscription"],
                "gym_membership": revenue_data.source_split["gym_membership"],
                "ai_credits": revenue_data.source_split.get("ai_credits", 0),
                "ai_diet_coach": revenue_data.source_split.get("ai_diet_coach", 0)
            },
            "sourceSplitRupees": {
                "daily_pass": revenue_data.source_split_rupees.get("daily_pass", 0),
                "sessions": revenue_data.source_split_rupees.get("sessions", 0),
                "fittbot_subscription": revenue_data.source_split_rupees.get("fittbot_subscription", 0),
                "gym_membership": revenue_data.source_split_rupees.get("gym_membership", 0),
                "ai_credits": revenue_data.source_split_rupees.get("ai_credits", 0),
                "ai_diet_coach": revenue_data.source_split_rupees.get("ai_diet_coach", 0)
            },
            "revenueOverTime": [
                {
                    "date": point.date,
                    "revenue": point.revenue
                }
                for point in revenue_data.daily_revenue
            ],
            "gymBreakdown": [
                {
                    "gym_id": point.gym_id,
                    "gym_name": point.gym_name,
                    "revenue": point.revenue
                }
                for point in revenue_data.gym_breakdown
            ],
            "filters": {
                "startDate": start_date_obj.isoformat(),
                "endDate": end_date_obj.isoformat(),
                "source": source or "all",
                "gymId": gym_id or "all"
            }
        }

        return {
            "success": True,
            "data": analytics_data,
            "message": "Revenue analytics fetched successfully"
        }

    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"Invalid date format: {str(e)}")
    except Exception as e:
        print(f"[DASHBOARD] Error in revenue-analytics: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/webinar-registrations-count")
async def get_webinar_registrations_count(db: AsyncSession = Depends(get_async_db)):
    """
    Get total count of webinar registrations from nutrition.webinar_registrations table.
    """
    try:
        from app.models.nutrition_models import WebinarRegistration
        stmt = select(func.count()).select_from(WebinarRegistration)
        result = await db.execute(stmt)
        total = result.scalar() or 0
        return {"success": True, "data": {"total": total}}
    except Exception as e:
        print(f"[DASHBOARD] Error in webinar-registrations-count: {str(e)}")
        return {"success": True, "data": {"total": 0}}

@router.get("/webinar-registrations")
async def get_webinar_registrations(db: AsyncSession = Depends(get_async_db)):
    """
    Get all webinar registrations from nutrition.webinar_registrations table.
    """
    try:
        from app.models.nutrition_models import WebinarRegistration
        stmt = select(WebinarRegistration).order_by(desc(WebinarRegistration.created_at))
        result = await db.execute(stmt)
        registrations = result.scalars().all()
        return {
            "success": True,
            "data": [
                {
                    "id": r.id,
                    "name": r.name,
                    "mobile_number": r.mobile_number,
                    "gender": r.gender,
                    "location": r.location,
                    "aim": r.aim,
                    "created_at": r.created_at.isoformat() if r.created_at else None,
                }
                for r in registrations
            ]
        }
    except Exception as e:
        print(f"[DASHBOARD] Error in webinar-registrations: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/recurring-subscribers")
async def get_recurring_subscribers(db: AsyncSession = Depends(get_async_db)):
    """
    Get count of recurring subscribers.
    NEW LOGIC: Returns clients who have purchased Nutritionist Plan more than once.
    Queries payments.payments where payment_metadata['flow'] = 'nutrition_purchase_googleplay' or 'nutrition_package_razorpay'
    """
    try:
        from collections import defaultdict

        # Dictionary to store nutritionist plan purchase count per customer_id
        customer_purchase_count = defaultdict(int)

        # NEW LOGIC: Query payments table for nutritionist plan purchases
        # Filters:
        # - payment_metadata['flow'] = 'nutrition_purchase_googleplay' or 'nutrition_package_razorpay'
        # - status = 'captured'
        payment_stmt = (
            select(Payment.customer_id)
            .join(
                NutritionEligibility,
                cast(Payment.order_id, String) == NutritionEligibility.source_id
            )
            .where(NutritionEligibility.source_type == "fymble_purchase")
            .where(Payment.status == "captured")
            .where(or_(
                func.json_extract(Payment.payment_metadata, '$.flow') == 'nutrition_purchase_googleplay',
                func.json_extract(Payment.payment_metadata, '$.flow') == 'nutrition_package_razorpay',
                func.json_extract(Payment.payment_metadata, '$.flow') == 'basic_nutrition_plan',
                func.json_extract(Payment.payment_metadata, '$.flow') == 'expert_nutrition_plan',
                func.json_extract(Payment.payment_metadata, '$.flow') == 'elite_nutrition_plan'
            ))
        )

        payment_result = await db.execute(payment_stmt)
        payments = payment_result.all()

        # Count each payment as a nutritionist plan purchase
        for payment in payments:
            customer_purchase_count[payment.customer_id] += 1

        # Filter only customers with more than 1 nutritionist plan purchase
        recurring_customer_ids = [
            customer_id for customer_id, count in customer_purchase_count.items()
            if count > 1
        ]

        # Convert string customer_ids to integers and check which ones exist in Client table
        customer_ids_int = []
        for customer_id_str in recurring_customer_ids:
            try:
                customer_ids_int.append(int(customer_id_str))
            except (ValueError, TypeError):
                pass

        # Fetch only valid client IDs that exist in Client table and are not excluded test contacts
        valid_count = 0
        if customer_ids_int:
            EXCLUDED_CONTACTS = ["7373675762", "9486987082", "8667458723", "9840633149", "8667427956", "8667488723", "7975847236"]
            stmt = (
                select(func.count())
                .select_from(Client)
                .where(
                    Client.client_id.in_(customer_ids_int),
                    ~Client.contact.in_(EXCLUDED_CONTACTS)
                )
            )
            result = await db.execute(stmt)
            valid_count = result.scalar() or 0

        return {
            "success": True,
            "data": {
                "total": valid_count
            },
            "message": "Recurring subscribers fetched successfully"
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching recurring subscribers: {str(e)}")


@router.get("/recurring-subscribers/details")
async def get_recurring_subscribers_details(
    page: int = 1,
    limit: int = 10,
    search: str = None,
    db: AsyncSession = Depends(get_async_db)
):
    """
    Get detailed list of recurring subscribers with pagination.
    Returns clients who have purchased Nutritionist Plan (Fittbot subscription) more than once.
    NEW LOGIC: Query payments.payments where payment_metadata['flow'] = 'nutrition_purchase_googleplay' or 'nutrition_package_razorpay'
    """
    try:
        from collections import defaultdict

        # Dictionary to store subscription info per customer_id
        customer_subscriptions = defaultdict(lambda: {"count": 0, "payments": []})

        # NEW LOGIC: Query payments table for nutritionist plan purchases
        # Filters:
        # - payment_metadata['flow'] = 'nutrition_purchase_googleplay' or 'nutrition_package_razorpay'
        # - status = 'captured'
        # Extract: customer_id, id, amount_minor, captured_at
        payment_stmt = (
            select(Payment.customer_id, Payment.id, Payment.amount_minor, Payment.captured_at)
            .join(
                NutritionEligibility,
                cast(Payment.order_id, String) == NutritionEligibility.source_id
            )
            .where(NutritionEligibility.source_type == "fymble_purchase")
            .where(Payment.status == "captured")
            .where(or_(
                func.json_extract(Payment.payment_metadata, '$.flow') == 'nutrition_purchase_googleplay',
                func.json_extract(Payment.payment_metadata, '$.flow') == 'nutrition_package_razorpay',
                func.json_extract(Payment.payment_metadata, '$.flow') == 'basic_nutrition_plan',
                func.json_extract(Payment.payment_metadata, '$.flow') == 'expert_nutrition_plan',
                func.json_extract(Payment.payment_metadata, '$.flow') == 'elite_nutrition_plan'
            ))
        )

        payment_result = await db.execute(payment_stmt)
        payments = payment_result.all()

        # Count each payment as a nutritionist plan purchase
        for payment in payments:
            customer_subscriptions[payment.customer_id]["count"] += 1
            customer_subscriptions[payment.customer_id]["payments"].append({
                "id": payment.id,
                "date": payment.captured_at.isoformat() if payment.captured_at else None,
                "amount": payment.amount_minor
            })

        # Filter only customers with more than 1 nutritionist plan purchase
        recurring_customers = {
            customer_id: data
            for customer_id, data in customer_subscriptions.items()
            if data["count"] > 1
        }

        # Get customer IDs and convert to integers for matching with Client table
        # The customer_id from payments is string, but Client.client_id is integer
        customer_ids_int = []
        customer_id_mapping = {}  # Maps int -> string (original customer_id)

        for customer_id_str in recurring_customers.keys():
            try:
                customer_id_int = int(customer_id_str)
                customer_ids_int.append(customer_id_int)
                customer_id_mapping[customer_id_int] = customer_id_str
            except (ValueError, TypeError):
                # Skip if customer_id cannot be converted to int
                pass

        # First, fetch all valid clients to get accurate count (excluding test contacts)
        valid_client_ids = []
        if customer_ids_int:
            EXCLUDED_CONTACTS = ["7373675762", "9486987082", "8667458723", "9840633149", "8667427956", "8667488723", "7975847236"]
            stmt = (
                select(Client.client_id)
                .where(
                    Client.client_id.in_(customer_ids_int),
                    ~Client.contact.in_(EXCLUDED_CONTACTS)
                )
            )
            result = await db.execute(stmt)
            valid_client_ids = [row[0] for row in result.all()]

        # Filter to only include customers that exist in Client table
        customer_ids_int = [cid for cid in customer_ids_int if cid in valid_client_ids]

        # Apply search filter if provided (search on ID, name, or contact)
        if search:
            # Fetch clients for filtering
            if customer_ids_int:
                stmt = select(Client).where(Client.client_id.in_(customer_ids_int))
                # Add search conditions
                search_pattern = f"%{search}%"
                stmt = stmt.where(
                    (Client.client_id.like(search_pattern)) |
                    (Client.name.like(search_pattern)) |
                    (Client.contact.like(search_pattern))
                )
                result = await db.execute(stmt)
                matching_clients = result.scalars().all()
                customer_ids_int = [client.client_id for client in matching_clients]

        # Get total count for pagination
        total_count = len(customer_ids_int)

        # Apply pagination
        start_idx = (page - 1) * limit
        end_idx = start_idx + limit
        paginated_customer_ids_int = customer_ids_int[start_idx:end_idx]

        # Fetch client details for the paginated customers
        subscribers_data = []
        if paginated_customer_ids_int:
            stmt = select(Client).where(Client.client_id.in_(paginated_customer_ids_int))
            result = await db.execute(stmt)
            clients = result.scalars().all()

            # Create a mapping of client_id to client data
            clients_map = {client.client_id: client for client in clients}

            for customer_id_int in paginated_customer_ids_int:
                client = clients_map.get(customer_id_int)
                if client:
                    # Get the original string customer_id for looking up subscriptions
                    customer_id_str = customer_id_mapping[customer_id_int]
                    subscriptions = recurring_customers[customer_id_str]

                    # Calculate total amount spent (using amount_minor from payments table)
                    total_amount = sum(
                        p["amount"] for p in subscriptions["payments"]
                    )

                    # Get first and last subscription date using captured_at
                    # Select the earliest entry for first_subscription
                    all_dates = [p["date"] for p in subscriptions["payments"] if p["date"]]
                    all_dates.sort()

                    subscribers_data.append({
                        "customer_id": customer_id_int,  # Use integer for display
                        "name": client.name or "N/A",
                        "contact": client.contact or "N/A",
                        "subscription_count": subscriptions["count"],
                        "total_amount": total_amount,
                        "first_subscription": all_dates[0] if all_dates else None,  # Earliest captured_at
                        "last_subscription": all_dates[-1] if all_dates else None,   # Latest captured_at
                    })

        return {
            "success": True,
            "data": {
                "subscribers": subscribers_data,
                "pagination": {
                    "page": page,
                    "limit": limit,
                    "total": total_count,
                    "total_pages": (total_count + limit - 1) // limit
                }
            },
            "message": "Recurring subscribers details fetched successfully"
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching recurring subscribers details: {str(e)}")


@router.get("/gymmate/profiles")
async def get_gymmate_profiles(
    page: int = 1,
    limit: int = 10,
    search: str = None,
    db: AsyncSession = Depends(get_async_db)
):
    """
    Get detailed list of GymMate profiles who have completed onboarding.
    Queries gym_mate.profile joined with clients table to get client details.
    """
    try:
        # Construct the base query selecting GymMateProfile and Client
        stmt = (
            select(GymMateProfile, Client.name, Client.contact, Client.email)
            .outerjoin(Client, GymMateProfile.client_id == Client.client_id)
            .where(GymMateProfile.onboarding_completed == True)
        )
        
        # Apply search filter if provided
        if search:
            search_pattern = f"%{search}%"
            stmt = stmt.where(
                or_(
                    Client.name.like(search_pattern),
                    Client.contact.like(search_pattern),
                    Client.email.like(search_pattern),
                    GymMateProfile.primary_goal.like(search_pattern),
                    GymMateProfile.gym_personality.like(search_pattern)
                )
            )
            
        # Get total count for pagination
        count_stmt = select(func.count()).select_from(stmt.subquery())
        count_result = await db.execute(count_stmt)
        total_count = count_result.scalar() or 0
        
        # Add pagination and sorting (latest created_at first)
        stmt = stmt.order_by(GymMateProfile.created_at.desc())
        stmt = stmt.offset((page - 1) * limit).limit(limit)
        
        result = await db.execute(stmt)
        rows = result.all()
        
        profiles_data = []
        for row in rows:
            profile = row[0]
            client_name = row[1]
            client_contact = row[2]
            client_email = row[3]
            
            profiles_data.append({
                "id": profile.id,
                "client_id": profile.client_id,
                "name": client_name or "N/A",
                "contact": client_contact or "N/A",
                "email": client_email or "N/A",
                "primary_goal": profile.primary_goal,
                "activity_interests": profile.activity_interests,
                "preferred_timing": profile.preferred_timing,
                "gym_personality": profile.gym_personality,
                "bio": profile.bio,
                "created_at": profile.created_at.isoformat() if profile.created_at else None,
                "updated_at": profile.updated_at.isoformat() if profile.updated_at else None
            })
            
        return {
            "success": True,
            "data": {
                "profiles": profiles_data,
                "pagination": {
                    "page": page,
                    "limit": limit,
                    "total": total_count,
                    "total_pages": (total_count + limit - 1) // limit
                }
            },
            "message": "GymMate profiles fetched successfully"
        }
    except Exception as e:
        print(f"[DASHBOARD] Error fetching GymMate profiles: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error fetching GymMate profiles: {str(e)}")


@router.get("/support-tickets-list")
async def get_support_tickets_list(
    source: str = Query(None, description="Filter by source: 'Fittbot' for client, 'Fittbot Business' for gym"),
    page: int = Query(1, ge=1, description="Page number"),
    limit: int = Query(10, ge=1, le=100, description="Items per page"),
    status: str = Query(None, description="Filter by status: 'all', 'resolved', 'unresolved', or 'follow_up'"),
    search: str = Query(None, description="Search by token or email"),
    priority: Optional[str] = Query(None, description="Filter by priority: 'high', 'medium', 'low', 'other'"),
    db: AsyncSession = Depends(get_async_db)
):

    import math
    try:

        if source == "Fittbot Business":
            # Query gym owner support tokens
            base_model = OwnerToken
            query = select(
                OwnerToken.id,
                OwnerToken.token.label('ticket_id'),
                literal("Fittbot Business").label('source'),
                Gym.name.label('name'),
                OwnerToken.email,
                OwnerToken.subject,
                OwnerToken.issue,
                OwnerToken.followed_up,
                OwnerToken.resolved,
                OwnerToken.comments,
                OwnerToken.created_at,
                OwnerToken.resolved_at,
                Employees.name.label('assigned_to'),
            ).outerjoin(
                Gym, OwnerToken.gym_id == Gym.gym_id
            ).outerjoin(
                TicketAssignment,
                (TicketAssignment.ticket_id == OwnerToken.id) & (TicketAssignment.ticket_source == "Fittbot Business") & (TicketAssignment.status == "active")
            ).outerjoin(
                Employees, Employees.id == TicketAssignment.employee_id
            )
        elif source == "Fittbot":
            # Query client support tokens
            base_model = ClientToken
            query = select(
                ClientToken.id,
                ClientToken.token.label('ticket_id'),
                literal("Fittbot").label('source'),
                Client.name.label('name'),
                ClientToken.email,
                ClientToken.subject,
                ClientToken.issue,
                ClientToken.followed_up,
                ClientToken.resolved,
                ClientToken.comments,
                ClientToken.created_at,
                ClientToken.resolved_at,
                Employees.name.label('assigned_to'),
            ).outerjoin(
                Client, ClientToken.client_id == Client.client_id
            ).outerjoin(
                TicketAssignment,
                (TicketAssignment.ticket_id == ClientToken.id) & (TicketAssignment.ticket_source == "Fittbot") & (TicketAssignment.status == "active")
            ).outerjoin(
                Employees, Employees.id == TicketAssignment.employee_id
            )
        else:
            raise HTTPException(status_code=400, detail="Invalid source. Use 'Fittbot' or 'Fittbot Business'")

        # Normalize status filter
        status_norm = (status or "").lower()
        if status_norm == "resolved":
            query = query.filter(base_model.resolved == True)
        elif status_norm in ("pending", "unresolved"):
            query = query.filter(base_model.resolved == False)
        elif status_norm in ("follow up", "follow_up"):
            query = query.filter(base_model.followed_up == True, base_model.resolved == False)
        elif status_norm == "all" or not status_norm:
            pass

        # Priority Lists based on source
        if source == "Fittbot Business":
            HIGH_PRIORITY_SUBJECTS = [
                "Unable to log in / OTP not received",
                "Payment not reflecting after collection",
                "Payment not received for Daily Pass/ Membership",
                "Daily Pass/Fitness Class/Membership QR scan not working",
                "Membership plan not activating after payment",
                "No Cost EMI not processing correctly",
                "Gym switching not working (multi-gym owners)",
                "Member fee incorrectly charged or duplicated",
                "Problem related to listing of my gym in Fymble App"
            ]

            MEDIUM_PRIORITY_SUBJECTS = [
                "Unable to create or edit membership plans",
                "Daily Pass Bookings not showing in calendar",
                "Fitness class/Membership bookings not showing up",
                "Fitness Class slot availability showing incorrect info",
                "Trainer unable to log in",
                "Unable to edit Daily Pass/ Fitness Class Price"
            ]

            LOW_PRIORITY_SUBJECTS = [
                "How to set up Daily Pass pricing",
                "How to create a fitness class",
                "How to import clients from another software",
                "How to add a trainer account",
                "Request to change gym name/details",
                "UI display issue / layout not aligned",
                "How to set up No Cost EMI",
                "Feature request — new functionality suggestion"
            ]
        else:
            HIGH_PRIORITY_SUBJECTS = [
                "Payment Failed but Amount Debited",
                "Refund Not Received",
                "Membership Purchased but Not Activated",
                "Daily Gym Pass Purchased but Entry Denied",
                "Nutrition Consultation Paid but Not Scheduled",
                "Nutrition Plan Access Issues",
                "Wrong Billing / Double Charge",
                "Account Hacked / Unauthorized Transactions",
                "OTP Failure",
                "Booked But Gym is Closed"
            ]

            MEDIUM_PRIORITY_SUBJECTS = [
                "Gym Membership Query",
                "Nutritionist Rescheduling",
                "Gym Mate Chat Issues",
                "XP Rewards Not Credited",
                "Referral Bonus",
                "Diet Tracker Sync Problems",
                "Workout Tracking Errors",
                "Food Scanner Incorrect Results",
                "App Crashes / Login Problems"
            ]

            LOW_PRIORITY_SUBJECTS = [
                "General Information Requests",
                "Feature Suggestions",
                "New Gym Request",
                "New Fitness Class Request",
                "Reward Program Questions",
                "How to Use Gym Mate",
                "Profile Update Assistance",
                "Feedback & Reviews",
                "Promotional Offer Queries"
            ]

        # Apply priority filter to query
        if priority:
            p_val = priority.lower().strip()
            if p_val == "high":
                query = query.filter(base_model.subject.in_(HIGH_PRIORITY_SUBJECTS))
            elif p_val == "medium":
                query = query.filter(base_model.subject.in_(MEDIUM_PRIORITY_SUBJECTS))
            elif p_val == "low":
                query = query.filter(base_model.subject.in_(LOW_PRIORITY_SUBJECTS))
            elif p_val == "other":
                query = query.filter(
                    ~base_model.subject.in_(HIGH_PRIORITY_SUBJECTS + MEDIUM_PRIORITY_SUBJECTS + LOW_PRIORITY_SUBJECTS) |
                    base_model.subject.is_(None)
                )

        # Apply search filter
        if search:
            search_term = f"%{search.lower()}%"
            query = query.filter(
                or_(
                    func.lower(base_model.token).like(search_term),
                    func.lower(base_model.email).like(search_term)
                )
            )

        # Order by created_at descending
        query = query.order_by(base_model.created_at.desc())

        # Get total count
        count_query = select(func.count()).select_from(base_model)
        if status_norm == "resolved":
            count_query = count_query.filter(base_model.resolved == True)
        elif status_norm in ("pending", "unresolved"):
            count_query = count_query.filter(base_model.resolved == False)
        elif status_norm in ("follow up", "follow_up"):
            count_query = count_query.filter(base_model.followed_up == True, base_model.resolved == False)
        
        # Apply priority filter to count_query
        if priority:
            p_val = priority.lower().strip()
            if p_val == "high":
                count_query = count_query.filter(base_model.subject.in_(HIGH_PRIORITY_SUBJECTS))
            elif p_val == "medium":
                count_query = count_query.filter(base_model.subject.in_(MEDIUM_PRIORITY_SUBJECTS))
            elif p_val == "low":
                count_query = count_query.filter(base_model.subject.in_(LOW_PRIORITY_SUBJECTS))
            elif p_val == "other":
                count_query = count_query.filter(
                    ~base_model.subject.in_(HIGH_PRIORITY_SUBJECTS + MEDIUM_PRIORITY_SUBJECTS + LOW_PRIORITY_SUBJECTS) |
                    base_model.subject.is_(None)
                )

        if search:
            search_term = f"%{search.lower()}%"
            count_query = count_query.filter(
                or_(
                    func.lower(base_model.token).like(search_term),
                    func.lower(base_model.email).like(search_term)
                )
            )

        count_result = await db.execute(count_query)
        total_count = count_result.scalar() or 0

        # Apply pagination
        offset = (page - 1) * limit
        query = query.offset(offset).limit(limit)

        # Execute query
        result = await db.execute(query)
        tickets_data = result.all()

        # Format response
        def map_status(followed_up, resolved):
           
            if resolved:
                return "Resolved"
            elif followed_up:
                return "Follow Up"
            else:
                return "Pending"

        tickets = []
        for row in tickets_data:
            tickets.append({
                "id": row.id,
                "ticket_id": row.ticket_id or f"ticket-{row.id}",
                "source": row.source,
                "name": row.name or "N/A",
                "email": row.email or "N/A",
                "subject": row.subject,
                "issue_type": row.subject,
                "issue": row.issue,
                "status": map_status(row.followed_up, row.resolved),
                "comments": row.comments,
                "created_at": row.created_at.isoformat() if row.created_at else None,
                "resolved_at": row.resolved_at.isoformat() if row.resolved_at else None,
                "assigned_to": row.assigned_to or None,
            })

        total_pages = (total_count + limit - 1) // limit if total_count > 0 else 1
        has_next = page < total_pages
        has_prev = page > 1

        return {
            "success": True,
            "message": "Tickets fetched successfully",
            "data": {
                "tickets": tickets,
                "total": total_count,
                "page": page,
                "limit": limit,
                "totalPages": total_pages,
                "hasNext": has_next,
                "hasPrev": has_prev
            }
        }
    except Exception as e:
        print(f"Error fetching support tickets: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error fetching support tickets: {str(e)}")

@router.get("/support-tickets-export")
async def export_support_tickets(
    source: str = Query(..., description="Fittbot or Fittbot Business"),
    start_date: str = Query(None),
    end_date: str = Query(None),
    db: AsyncSession = Depends(get_async_db)
):
    """
    Export support tickets to Excel with 4 sheets based on status.
    Applies date range filter.
    """
    try:
        import io
        import pandas as pd
        from fastapi.responses import StreamingResponse

        # Determine models based on source
        if source == "Fittbot Business":
            base_model = OwnerToken
            user_model = Gym
            join_col = OwnerToken.gym_id == Gym.gym_id
            ticket_source_val = "Fittbot Business"
        elif source == "Fittbot":
            base_model = ClientToken
            user_model = Client
            join_col = ClientToken.client_id == Client.client_id
            ticket_source_val = "Fittbot"
        else:
            raise HTTPException(status_code=400, detail="Invalid source. Use 'Fittbot' or 'Fittbot Business'")

        # Build base query
        stmt = select(
            base_model.token.label('ticket_id'),
            base_model.subject,
            base_model.issue.label('issue_description'),
            base_model.comments,
            base_model.followed_up,
            base_model.resolved,
            base_model.created_at,
            base_model.resolved_at,
            Employees.name.label('assigned_to')
        ).outerjoin(
            user_model, join_col
        ).outerjoin(
            TicketAssignment,
            (TicketAssignment.ticket_id == base_model.id) & (TicketAssignment.ticket_source == ticket_source_val) & (TicketAssignment.status == "active")
        ).outerjoin(
            Employees, Employees.id == TicketAssignment.employee_id
        )

        # Apply date filters
        if start_date:
            stmt = stmt.filter(func.date(base_model.created_at) >= start_date)
        if end_date:
            stmt = stmt.filter(func.date(base_model.created_at) <= end_date)

        # Order by created_at descending
        stmt = stmt.order_by(base_model.created_at.desc())

        # Execute query
        result = await db.execute(stmt)
        all_tickets = result.all()

        # Group data for sheets
        def get_status(followed_up, resolved):
            if resolved: return "Resolved"
            elif followed_up: return "Follow Up"
            else: return "Pending"

        data_all = []
        data_resolved = []
        data_unresolved = []
        data_followup = []

        for row in all_tickets:
            status = get_status(row.followed_up, row.resolved)
            item = {
                "Ticket ID": row.ticket_id,
                "Subject": row.subject or "N/A",
                "Status": status,
                "Assigned To": row.assigned_to or "N/A",
                "Issue Description": row.issue_description or "N/A",
                "Comments": row.comments or "N/A",
                "Created At": row.created_at.strftime("%Y-%m-%d %H:%M:%S") if row.created_at else "N/A",
                "Resolved At": row.resolved_at.strftime("%Y-%m-%d %H:%M:%S") if row.resolved_at else "N/A"
            }
            
            data_all.append(item)
            if status == "Resolved":
                data_resolved.append(item)
            elif status == "Pending":
                data_unresolved.append(item)
            elif status == "Follow Up":
                data_followup.append(item)

        # Create Excel in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Sheets configuration
            sheets = [
                ("All Tickets", data_all),
                ("Resolved", data_resolved),
                ("Unresolved", data_unresolved),
                ("Follow Up", data_followup)
            ]
            
            for sheet_name, data in sheets:
                df = pd.DataFrame(data)
                if df.empty:
                    # Create even if empty to satisfy the requirement of 4 sheets
                    df = pd.DataFrame(columns=["Ticket ID", "Subject", "Status", "Assigned To", "Issue Description", "Comments", "Created At", "Resolved At"])
                
                df.to_excel(writer, index=False, sheet_name=sheet_name)
                
                # Auto-adjust column widths and style header
                worksheet = writer.sheets[sheet_name]
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")  # Steel Blue
                header_font = Font(color="FFFFFF", bold=True)
                
                for idx, col in enumerate(df.columns, 1):
                    # Style header cell
                    cell = worksheet.cell(row=1, column=idx)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
                    
                    series = df[col]
                    max_len = max(
                        series.astype(str).map(len).max(),
                        len(str(col))
                    ) + 2
                    col_letter = worksheet.cell(row=1, column=idx).column_letter
                    worksheet.column_dimensions[col_letter].width = min(max_len, 60)

        output.seek(0)
        
        # Filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_source = source.replace(" ", "_")
        filename = f"Support_Tickets_{safe_source}_{timestamp}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        print(f"Error exporting support tickets: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error exporting support tickets: {str(e)}")
    

@router.get("/gym-ticket-detail")
async def get_gym_ticket_detail(
    ticket_id: str = Query(..., description="Ticket token string"),
    db: AsyncSession = Depends(get_async_db)
):
    """Get gym (owner) ticket details by token string"""
    try:
        query = select(
            OwnerToken.id,
            OwnerToken.token.label('ticket_id'),
            literal("Fittbot Business").label('source'),
            Gym.name.label('name'),
            OwnerToken.email,
            OwnerToken.subject,
            OwnerToken.issue,
            OwnerToken.followed_up,
            OwnerToken.resolved,
            OwnerToken.comments,
            OwnerToken.created_at
        ).outerjoin(
            Gym, OwnerToken.gym_id == Gym.gym_id
        ).filter(
            OwnerToken.token == ticket_id
        )

        result = await db.execute(query)
        ticket_data = result.first()

        if not ticket_data:
            raise HTTPException(status_code=404, detail="Ticket not found")

        def map_status(followed_up, resolved):
            if resolved:
                return "resolved"
            else:
                return "pending"

        ticket = {
            "id": ticket_data.id,
            "ticket_id": ticket_data.ticket_id or f"ticket-{ticket_data.id}",
            "source": ticket_data.source,
            "name": ticket_data.name or "N/A",
            "email": ticket_data.email or "N/A",
            "subject": ticket_data.subject,
            "issue": ticket_data.issue,
            "status": map_status(ticket_data.followed_up, ticket_data.resolved),
            "comments": ticket_data.comments,
            "created_at": ticket_data.created_at.isoformat() if ticket_data.created_at else None,
        }

        return {
            "success": True,
            "data": ticket,
            "message": "Ticket details fetched successfully"
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching ticket details: {str(e)}")



@router.get("/client-ticket-detail")
async def get_client_ticket_detail(
    ticket_id: str = Query(..., description="Ticket token string"),
    db: AsyncSession = Depends(get_async_db)
):
   
    import sys
    try:

        query = select(
            ClientToken.id,
            ClientToken.token.label('ticket_id'),
            literal("Fittbot").label('source'),
            Client.name.label('name'),
            ClientToken.email,
            ClientToken.subject,
            ClientToken.issue,
            ClientToken.followed_up,
            ClientToken.resolved,
            ClientToken.comments,
            ClientToken.created_at
        ).outerjoin(
            Client, ClientToken.client_id == Client.client_id
        ).filter(
            ClientToken.token == ticket_id
        )

        result = await db.execute(query)
        ticket_data = result.first()

        if not ticket_data:
            raise HTTPException(status_code=404, detail="Ticket not found")

        def map_status(followed_up, resolved):
           
            if resolved:
                return "Resolved"
            elif followed_up:
                return "Follow Up"
            else:
                return "Pending"

        ticket = {
            "id": ticket_data.id,
            "ticket_id": ticket_data.ticket_id or f"ticket-{ticket_data.id}",
            "source": ticket_data.source,
            "name": ticket_data.name or "N/A",
            "email": ticket_data.email or "N/A",
            "subject": ticket_data.subject,
            "issue": ticket_data.issue,
            "status": map_status(ticket_data.followed_up, ticket_data.resolved),
            "comments": ticket_data.comments,
            "created_at": ticket_data.created_at.isoformat() if ticket_data.created_at else None,
        }

        return {
            "success": True,
            "data": ticket,
            "message": "Ticket details fetched successfully"
        }

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc(file=sys.stderr)
        print(f"[CLIENT-TICKET-DETAIL-ERROR] {str(e)}", file=sys.stderr)
        raise HTTPException(status_code=500, detail=f"Error fetching ticket details: {str(e)}")

class MarkResolvedRequest(BaseModel):
    ticket_id: str


@router.post("/gym-ticket-resolve")
async def mark_gym_ticket_resolved(
    request: MarkResolvedRequest,
    db: AsyncSession = Depends(get_async_db)
):
    """Mark gym ticket as resolved"""
    try:
        query = select(OwnerToken).filter(OwnerToken.token == request.ticket_id)
        result = await db.execute(query)
        ticket = result.scalar_one_or_none()

        if not ticket:
            raise HTTPException(status_code=404, detail="Ticket not found")

        ticket.resolved = True
        ticket.followed_up = True
        ticket.updated_at = datetime.now()

        await db.commit()

        return {
            "success": True,
            "message": "Ticket marked as resolved"
        }

    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Error marking ticket as resolved: {str(e)}")


@router.post("/client-ticket-resolve")
async def mark_client_ticket_resolved(
    request: MarkResolvedRequest,
    db: AsyncSession = Depends(get_async_db)
):
    """Mark client ticket as resolved"""
    try:
        query = select(ClientToken).filter(ClientToken.token == request.ticket_id)
        result = await db.execute(query)
        ticket = result.scalar_one_or_none()

        if not ticket:
            raise HTTPException(status_code=404, detail="Ticket not found")

        ticket.resolved = True
        ticket.followed_up = True
        ticket.updated_at = datetime.now()

        await db.commit()

        return {
            "success": True,
            "message": "Ticket marked as resolved"
        }

    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Error marking ticket as resolved: {str(e)}")


class TicketFollowUpRequest(BaseModel):
    ticket_id: str
    source: str  # "client" or "owner"
    comment: Optional[str] = None
    status: Optional[str] = None  # "followup"


@router.post("/ticket_followup")
async def ticket_followup(
    request: TicketFollowUpRequest,
    db: AsyncSession = Depends(get_async_db)
):
    """
    Mark a ticket for follow-up or add comments to a ticket
    """
    import sys
    try:
        if not request.comment and not request.status:
            raise HTTPException(status_code=400, detail="At least one of comment or status must be provided")

        # Select the correct table based on source
        if request.source == "client":
            query = select(ClientToken).filter(ClientToken.token == request.ticket_id)
        elif request.source == "owner":
            query = select(OwnerToken).filter(OwnerToken.token == request.ticket_id)
        else:
            raise HTTPException(status_code=400, detail="Invalid source. Use 'client' or 'owner'")

        result = await db.execute(query)
        ticket = result.scalar_one_or_none()

        if not ticket:
            raise HTTPException(status_code=404, detail="Ticket not found")

        # Update status to followup if requested
        if request.status and request.status == "follow_up":
            ticket.followed_up = True
            ticket.resolved = False

        # Add comment if provided
        if request.comment:
            ticket.comments = request.comment

        ticket.updated_at = datetime.now()
        await db.commit()

        return {
            "success": True,
            "message": "Ticket updated successfully"
        }

    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        import traceback
        traceback.print_exc(file=sys.stderr)
        print(f"[TICKET-FOLLOWUP-ERROR] {str(e)}", file=sys.stderr)
        raise HTTPException(status_code=500, detail=f"Error updating ticket: {str(e)}")


@router.get("/purchase-analytics")
async def get_purchase_analytics(
    start_date: str = None,
    end_date: str = None,
    source: str = None,
    gym_id: int = None,
    location: str = None,
    db: AsyncSession = Depends(get_async_db)
):

    try:
        # Debug: Log the gym_id parameter
        import logging
        logging.info(f"Purchase analytics called with gym_id: {gym_id} (type: {type(gym_id)}), start_date: {start_date}, end_date: {end_date}, source: {source}")

        # Parse dates if provided
        if start_date:
            start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
        else:
            # Default to early date for overall data
            start_date_obj = datetime(1970, 1, 1).date()

        if end_date:
            end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
        else:
            # Default to today
            end_date_obj = datetime(2100, 1, 1).date()

        # Get client IDs filtered by location if location is provided
        location_client_ids = set()
        if location and location != "all":
            try:
                location_stmt = select(Client.client_id).where(Client.location == location)
                location_result = await db.execute(location_stmt)
                location_client_ids = set(row[0] for row in location_result.all())
            except Exception as e:
                location_client_ids = set()
        else:
            # Debug: Check how many purchases have NULL or invalid client_id
            try:
                dailypass_session = get_dailypass_session()

                # Count total DailyPass purchases
                total_dp = dailypass_session.query(func.count()).scalar()

                # Count DailyPass with NULL client_id
                null_client_dp = dailypass_session.query(func.count()).filter(DailyPass.client_id.is_(None)).scalar()

                # Count DailyPass with client_id NOT in Client table
                valid_client_ids = select(Client.client_id)
                not_in_client = dailypass_session.query(func.count()).filter(~DailyPass.client_id.in_(valid_client_ids)).scalar()

                dailypass_session.close()
            except Exception as e:
                pass

        # Initialize result structure
        category_breakdown = {
            "daily_pass": {"purchases": 0, "unique_users": 0, "purchases_over_time": []},
            "sessions": {"purchases": 0, "unique_users": 0, "purchases_over_time": []},
            "fittbot_subscription": {"purchases": 0, "unique_users": 0, "purchases_over_time": []},
            "ai_credits": {"purchases": 0, "unique_users": 0, "purchases_over_time": []},
            "ai_diet_coach": {"purchases": 0, "unique_users": 0, "purchases_over_time": []},
            "gym_membership": {"purchases": 0, "unique_users": 0, "purchases_over_time": []}
        }

        all_purchases_over_time = {}  # date -> total count
        gym_purchases = {}  # gym_id -> purchase count
        location_purchases = {}  # location -> purchase count

        # 1. DAILY PASS PURCHASES - Single aggregated query
        # GMV fix: INNER JOIN with Gym table to exclude orphaned DailyPass records (gym_id not in Gym table)
        if not source or source == "daily_pass":
            try:
                import logging
                logging.info(f"Daily pass purchases: gym_id={gym_id}, source={source}")
                dailypass_session = get_dailypass_session()

                # Build base query — JOIN Gym table to exclude records with no matching gym
                # Treated as exactly 1 purchase per pass transaction
                all_query = dailypass_session.query(
                    func.count(DailyPass.id).label('total_purchases'),
                    func.count(distinct(DailyPass.client_id)).label('total_unique_users')
                ).join(
                    Gym, Gym.gym_id == func.cast(DailyPass.gym_id, Integer), isouter=False
                ).filter(
                    func.date(DailyPass.created_at) >= start_date_obj,
                    func.date(DailyPass.created_at) <= end_date_obj,
                    DailyPass.gym_id != "1"
                )

                # Apply gym filter if provided
                if gym_id:
                    gym_id_str = str(gym_id)
                    all_query = all_query.filter(DailyPass.gym_id == gym_id_str)
                    logging.info(f"Daily pass query with gym filter: gym_id_str={gym_id_str}")

                # Apply location filter if provided
                if location_client_ids:
                    all_query = all_query.filter(DailyPass.client_id.in_(location_client_ids))

                # Get totals
                total_result = all_query.first()

                if total_result:
                    category_breakdown["daily_pass"]["purchases"] = total_result.total_purchases or 0
                    category_breakdown["daily_pass"]["unique_users"] = total_result.total_unique_users or 0

                # Build separate query for purchases over time (grouped by date) — same Gym JOIN
                # Treated as exactly 1 purchase per pass transaction
                base_query = dailypass_session.query(
                    func.date(DailyPass.created_at).label('purchase_date'),
                    func.count(DailyPass.id).label('purchase_count')
                ).join(
                    Gym, Gym.gym_id == func.cast(DailyPass.gym_id, Integer), isouter=False
                ).filter(
                    func.date(DailyPass.created_at) >= start_date_obj,
                    func.date(DailyPass.created_at) <= end_date_obj,
                    DailyPass.gym_id != "1"
                )

                # Apply gym filter if provided
                if gym_id:
                    base_query = base_query.filter(DailyPass.gym_id == str(gym_id))

                # Apply location filter if provided
                if location_client_ids:
                    base_query = base_query.filter(DailyPass.client_id.in_(location_client_ids))

                # Group by date and get results
                base_query = base_query.group_by(func.date(DailyPass.created_at))
                time_result = base_query.all()

                if time_result:
                    # Build purchases over time
                    for row in time_result:
                        date_key = row.purchase_date.isoformat() if row.purchase_date else None
                        if date_key:
                            if date_key not in all_purchases_over_time:
                                all_purchases_over_time[date_key] = 0
                            all_purchases_over_time[date_key] += row.purchase_count
                            category_breakdown["daily_pass"]["purchases_over_time"].append({
                                "date": date_key,
                                "purchases": row.purchase_count
                            })

             
                category_breakdown["daily_pass"]["purchases_over_time"].sort(key=lambda x: x["date"])

                # Get gym-wise purchases - ALWAYS get all gyms with purchases in timeframe/location
                # This ensures the filter dropdown remains populated even when a gym is selected.
                # Treated as exactly 1 purchase per pass transaction
                gym_query = dailypass_session.query(
                    DailyPass.gym_id,
                    func.count(DailyPass.id).label('purchase_count')
                ).join(
                    Gym, Gym.gym_id == func.cast(DailyPass.gym_id, Integer), isouter=False
                ).filter(
                    func.date(DailyPass.created_at) >= start_date_obj,
                    func.date(DailyPass.created_at) <= end_date_obj,
                    DailyPass.gym_id.isnot(None),
                    DailyPass.gym_id != "1"  # Exclude gym_id = 1
                )

                # Apply location filter if provided
                if location_client_ids:
                    gym_query = gym_query.filter(DailyPass.client_id.in_(location_client_ids))

                gym_query = gym_query.group_by(DailyPass.gym_id)
                gym_result = gym_query.all()

                for row in gym_result:
                    try:
                        gym_key = int(row.gym_id)
                        if gym_key not in gym_purchases:
                            gym_purchases[gym_key] = 0
                        gym_purchases[gym_key] += row.purchase_count
                    except (ValueError, TypeError):
                        pass
                
                if gym_id:
                    logging.info(f"Daily pass for filtered gym {gym_id}: {gym_purchases.get(int(gym_id), 0)} purchases")

                logging.info(f"Daily pass gym_purchases: {gym_purchases}")

                # Get location-wise purchases (only when no location filter is applied)
                # GMV fix: add Gym INNER JOIN so orphaned DailyPass records are excluded
                if not location or location == "all":
                    try:
                        # Get all DailyPass client_ids and their headcount-scaled purchases — with Gym JOIN to exclude orphans
                        daily_pass_records = dailypass_session.query(
                            DailyPass.client_id,
                            literal(1).label("purchase_count")
                        ).join(
                            Gym, Gym.gym_id == func.cast(DailyPass.gym_id, Integer), isouter=False
                        ).filter(
                            func.date(DailyPass.created_at) >= start_date_obj,
                            func.date(DailyPass.created_at) <= end_date_obj,
                            DailyPass.client_id.isnot(None),
                            DailyPass.gym_id != "1"
                        )

                        if gym_id:
                            daily_pass_records = daily_pass_records.filter(DailyPass.gym_id == str(gym_id))

                        daily_pass_records = daily_pass_records.all()

                        client_ids = list(set([str(r.client_id) for r in daily_pass_records if r.client_id]))

                        if client_ids:
                            client_location_stmt = select(Client.client_id, Client.location).where(
                                Client.client_id.in_(client_ids),
                                Client.location.isnot(None),
                                Client.location != ''
                            )
                            client_location_result = await db.execute(client_location_stmt)
                            client_locations = {str(row[0]): row[1] for row in client_location_result.all()}

                            for record in daily_pass_records:
                                if record.client_id and str(record.client_id) in client_locations:
                                    raw_loc = client_locations[str(record.client_id)]
                                    normalized_loc = raw_loc.strip().replace(' ', '_')
                                    if normalized_loc not in location_purchases:
                                        location_purchases[normalized_loc] = 0
                                    location_purchases[normalized_loc] += record.purchase_count

                    except Exception as e:
                        import logging
                        logging.error(f"Error getting location-wise daily pass purchases: {str(e)}")


                dailypass_session.close()
            except Exception as e:
                # Log error for debugging daily pass purchases
                import logging
                logging.error(f"Purchase analytics - Daily pass error: {str(e)}")
                pass

        # 2. SESSION PURCHASES - Use SessionPurchase (status='paid') + Gym INNER JOIN (matches GMV logic)
        if not source or source == "sessions":
            try:
                # GMV fix: SessionPurchase with status='paid' + INNER JOIN Gym (no orphans)
                # Treated as exactly 1 purchase per transaction
                total_stmt = (
                    select(
                        func.count(SessionPurchase.id).label('total_purchases'),
                        func.count(distinct(SessionPurchase.client_id)).label('total_unique_users')
                    )
                    .select_from(SessionPurchase)
                    .join(Gym, SessionPurchase.gym_id == Gym.gym_id)
                    .where(
                        SessionPurchase.status == "paid",
                        SessionPurchase.gym_id != 1,
                        func.date(SessionPurchase.created_at) >= start_date_obj,
                        func.date(SessionPurchase.created_at) <= end_date_obj,
                    )
                )

                # Apply gym filter if provided
                if gym_id:
                    total_stmt = total_stmt.where(SessionPurchase.gym_id == gym_id)

                # Apply location filter if provided
                if location_client_ids:
                    total_stmt = total_stmt.where(SessionPurchase.client_id.in_(location_client_ids))

                total_result = await db.execute(total_stmt)
                total_row = total_result.first()

                if total_row:
                    category_breakdown["sessions"]["purchases"] = total_row.total_purchases or 0
                    category_breakdown["sessions"]["unique_users"] = total_row.total_unique_users or 0

                # Then get purchases over time (grouped by date) — same model/filters
                # Treated as exactly 1 purchase per transaction in time-series
                time_stmt = (
                    select(
                        func.date(SessionPurchase.created_at).label('purchase_date'),
                        func.count(SessionPurchase.id).label('purchase_count')
                    )
                    .select_from(SessionPurchase)
                    .join(Gym, SessionPurchase.gym_id == Gym.gym_id)
                    .where(
                        SessionPurchase.status == "paid",
                        SessionPurchase.gym_id != 1,
                        func.date(SessionPurchase.created_at) >= start_date_obj,
                        func.date(SessionPurchase.created_at) <= end_date_obj,
                    )
                )

                # Apply gym filter if provided
                if gym_id:
                    time_stmt = time_stmt.where(SessionPurchase.gym_id == gym_id)

                # Apply location filter if provided
                if location_client_ids:
                    time_stmt = time_stmt.where(SessionPurchase.client_id.in_(location_client_ids))

                # Group by date and execute
                time_stmt = time_stmt.group_by(func.date(SessionPurchase.created_at))
                time_result = await db.execute(time_stmt)
                session_results = time_result.all()

                if session_results:
                    # Build purchases over time
                    for row in session_results:
                        date_key = row.purchase_date.isoformat() if row.purchase_date else None
                        if date_key:
                            if date_key not in all_purchases_over_time:
                                all_purchases_over_time[date_key] = 0
                            all_purchases_over_time[date_key] += row.purchase_count
                            category_breakdown["sessions"]["purchases_over_time"].append({
                                "date": date_key,
                                "purchases": row.purchase_count
                            })

                # Sort purchases over time by date
                category_breakdown["sessions"]["purchases_over_time"].sort(key=lambda x: x["date"])

                # Get gym-wise purchases — ALWAYS get all gyms (for filter list stability)
                gym_wise_stmt = (
                    select(
                        SessionPurchase.gym_id,
                        func.count().label('purchase_count')
                    )
                    .select_from(SessionPurchase)
                    .join(Gym, SessionPurchase.gym_id == Gym.gym_id)
                    .where(
                        SessionPurchase.status == "paid",
                        func.date(SessionPurchase.created_at) >= start_date_obj,
                        func.date(SessionPurchase.created_at) <= end_date_obj,
                        SessionPurchase.gym_id.isnot(None),
                        SessionPurchase.gym_id != 1
                    )
                )

                if location_client_ids:
                    gym_wise_stmt = gym_wise_stmt.where(SessionPurchase.client_id.in_(location_client_ids))

                gym_wise_stmt = gym_wise_stmt.group_by(SessionPurchase.gym_id)
                gym_wise_result = await db.execute(gym_wise_stmt)
                gym_wise_rows = gym_wise_result.all()

                for row in gym_wise_rows:
                    if row.gym_id:
                        if row.gym_id not in gym_purchases:
                            gym_purchases[row.gym_id] = 0
                        gym_purchases[row.gym_id] += row.purchase_count
                
                if gym_id:
                    logging.info(f"Sessions for filtered gym {gym_id}: {gym_purchases.get(int(gym_id), 0)} purchases")

                # Get location-wise purchases (only when no location filter is applied)
                if not location or location == "all":
                    try:
                        session_records_stmt = (
                            select(SessionPurchase.client_id)
                            .select_from(SessionPurchase)
                            .join(Gym, SessionPurchase.gym_id == Gym.gym_id)
                            .where(
                                SessionPurchase.status == "paid",
                                func.date(SessionPurchase.created_at) >= start_date_obj,
                                func.date(SessionPurchase.created_at) <= end_date_obj,
                                SessionPurchase.client_id.isnot(None),
                                SessionPurchase.gym_id != 1
                            )
                        )

                        if gym_id:
                            session_records_stmt = session_records_stmt.where(SessionPurchase.gym_id == gym_id)

                        if location_client_ids:
                            session_records_stmt = session_records_stmt.where(SessionPurchase.client_id.in_(location_client_ids))

                        session_records_result = await db.execute(session_records_stmt)
                        session_records = session_records_result.all()

                        # Get unique client_ids
                        client_ids = list(set([str(row[0]) for row in session_records]))

                        if client_ids:
                            # Query Client table to get locations
                            client_location_stmt = select(Client.client_id, Client.location).where(
                                Client.client_id.in_(client_ids),
                                Client.location.isnot(None),
                                Client.location != ''
                            )
                            client_location_result = await db.execute(client_location_stmt)
                            client_locations = {str(row[0]): row[1] for row in client_location_result.all()}

                            # Count purchases per location (normalize location names)
                            for row in session_records:
                                if row[0] and str(row[0]) in client_locations:
                                    raw_loc = client_locations[str(row[0])]
                                    # Normalize location: trim whitespace and replace spaces with underscores
                                    normalized_loc = raw_loc.strip().replace(' ', '_')
                                    if normalized_loc not in location_purchases:
                                        location_purchases[normalized_loc] = 0
                                    location_purchases[normalized_loc] += 1

                    except Exception as e:
                        import logging
                        logging.error(f"Error getting location-wise session purchases: {str(e)}")

            except Exception:
                pass

        # 3. NUTRITIONIST PLAN (FITTBOT SUBSCRIPTION) PURCHASES
        # GMV fix: added excluded contacts filter to match gmv-summary logic
        # NOTE: Skip when gym filter is applied (not gym-specific purchases)
        EXCLUDED_CONTACTS = ["7373675762", "9486987082", "8667458723", "9840633149", "8667427956", "8667488723", "7975847236"]
        if (not source or source == "fittbot_subscription") and not gym_id:
            try:
                nutritionist_stmt = (
                    select(
                        func.date(Payment.captured_at).label('purchase_date'),
                        func.count().label('purchase_count'),
                        func.count(distinct(Payment.customer_id)).label('unique_users')
                    )
                    .select_from(Payment)
                    .outerjoin(Client, Payment.customer_id == Client.client_id)
                    .join(
                        NutritionEligibility,
                        cast(Payment.order_id, String) == NutritionEligibility.source_id
                    )
                    .where(NutritionEligibility.source_type == "fymble_purchase")
                    .where(Payment.status == "captured")
                    .where(or_(
                        func.json_extract(Payment.payment_metadata, '$.flow') == 'nutrition_purchase_googleplay',
                        func.json_extract(Payment.payment_metadata, '$.flow') == 'nutrition_package_razorpay',
                        func.json_extract(Payment.payment_metadata, '$.flow') == 'basic_nutrition_plan',
                        func.json_extract(Payment.payment_metadata, '$.flow') == 'expert_nutrition_plan',
                        func.json_extract(Payment.payment_metadata, '$.flow') == 'elite_nutrition_plan'
                    ))
                    .where(func.date(Payment.captured_at) >= start_date_obj)
                    .where(func.date(Payment.captured_at) <= end_date_obj)
                    .where(~Client.contact.in_(EXCLUDED_CONTACTS))
                )

                if location_client_ids:
                    location_customer_ids = {str(cid) for cid in location_client_ids}
                    nutritionist_stmt = nutritionist_stmt.where(Payment.customer_id.in_(location_customer_ids))

                nutritionist_stmt = nutritionist_stmt.group_by(func.date(Payment.captured_at))

                result = await db.execute(nutritionist_stmt)
                nutritionist_results = result.all()

                # Get total unique users across all dates
                unique_users_stmt = (
                    select(func.count(distinct(Payment.customer_id)))
                    .select_from(Payment)
                    .outerjoin(Client, Payment.customer_id == Client.client_id)
                    .join(
                        NutritionEligibility,
                        cast(Payment.order_id, String) == NutritionEligibility.source_id
                    )
                    .where(NutritionEligibility.source_type == "fymble_purchase")
                    .where(Payment.status == "captured")
                    .where(or_(
                        func.json_extract(Payment.payment_metadata, '$.flow') == 'nutrition_purchase_googleplay',
                        func.json_extract(Payment.payment_metadata, '$.flow') == 'nutrition_package_razorpay',
                        func.json_extract(Payment.payment_metadata, '$.flow') == 'basic_nutrition_plan',
                        func.json_extract(Payment.payment_metadata, '$.flow') == 'expert_nutrition_plan',
                        func.json_extract(Payment.payment_metadata, '$.flow') == 'elite_nutrition_plan'
                    ))
                    .where(func.date(Payment.captured_at) >= start_date_obj)
                    .where(func.date(Payment.captured_at) <= end_date_obj)
                    .where(~Client.contact.in_(EXCLUDED_CONTACTS))
                )

                if location_client_ids:
                    location_customer_ids = {str(cid) for cid in location_client_ids}
                    unique_users_stmt = unique_users_stmt.where(Payment.customer_id.in_(location_customer_ids))

                unique_result = await db.execute(unique_users_stmt)
                unique_users_count = unique_result.scalar() or 0

                # Calculate totals and build purchases over time
                total_purchases = 0
                for row in nutritionist_results:
                    date_key = row.purchase_date.isoformat() if row.purchase_date else None
                    if date_key:
                        total_purchases += row.purchase_count
                        if date_key not in all_purchases_over_time:
                            all_purchases_over_time[date_key] = 0
                        all_purchases_over_time[date_key] += row.purchase_count
                        category_breakdown["fittbot_subscription"]["purchases_over_time"].append({
                            "date": date_key,
                            "purchases": row.purchase_count
                        })

                category_breakdown["fittbot_subscription"]["purchases"] = total_purchases
                category_breakdown["fittbot_subscription"]["unique_users"] = unique_users_count

            except Exception as e:
                import logging
                logging.error(f"Purchase analytics - Nutritionist Plan error: {str(e)}")
                pass

        # 3.5. AI CREDITS PURCHASES
        # GMV fix: added excluded contacts filter to match gmv-summary logic
        # NOTE: Skip when gym filter is applied (not gym-specific purchases)
        if (not source or source == "ai_credits") and not gym_id:
            try:
                ai_credits_stmt = (
                    select(
                        func.date(Payment.captured_at).label('purchase_date'),
                        func.count().label('purchase_count'),
                        func.count(distinct(Payment.customer_id)).label('unique_users')
                    )
                    .select_from(Payment)
                    .outerjoin(Client, Payment.customer_id == Client.client_id)
                    .where(Payment.status == "captured")
                    .where(or_(
                        func.json_extract(Payment.payment_metadata, '$.flow') == 'food_scanner_credits',
                        func.json_extract(Payment.payment_metadata, '$.flow') == 'food_scanner_credits_razorpay'
                    ))
                    .where(func.date(Payment.captured_at) >= start_date_obj)
                    .where(func.date(Payment.captured_at) <= end_date_obj)
                    .where(~Client.contact.in_(EXCLUDED_CONTACTS))
                )

                if location_client_ids:
                    location_customer_ids = {str(cid) for cid in location_client_ids}
                    ai_credits_stmt = ai_credits_stmt.where(Payment.customer_id.in_(location_customer_ids))

                ai_credits_stmt = ai_credits_stmt.group_by(func.date(Payment.captured_at))

                result = await db.execute(ai_credits_stmt)
                ai_credits_results = result.all()

                # Get total unique users across all dates
                ai_credits_unique_users_stmt = (
                    select(func.count(distinct(Payment.customer_id)))
                    .select_from(Payment)
                    .outerjoin(Client, Payment.customer_id == Client.client_id)
                    .where(Payment.status == "captured")
                    .where(or_(
                        func.json_extract(Payment.payment_metadata, '$.flow') == 'food_scanner_credits',
                        func.json_extract(Payment.payment_metadata, '$.flow') == 'food_scanner_credits_razorpay'
                    ))
                    .where(func.date(Payment.captured_at) >= start_date_obj)
                    .where(func.date(Payment.captured_at) <= end_date_obj)
                    .where(~Client.contact.in_(EXCLUDED_CONTACTS))
                )

                if location_client_ids:
                    location_customer_ids = {str(cid) for cid in location_client_ids}
                    ai_credits_unique_users_stmt = ai_credits_unique_users_stmt.where(Payment.customer_id.in_(location_customer_ids))

                ai_credits_unique_result = await db.execute(ai_credits_unique_users_stmt)
                ai_credits_unique_users_count = ai_credits_unique_result.scalar() or 0

                # Compute AI Credits Total Purchases exactly like compute_gmv_totals
                ai_total_stmt = (
                    select(func.count(Payment.id))
                    .select_from(Payment)
                    .outerjoin(Client, Payment.customer_id == Client.client_id)
                    .where(Payment.status == "captured")
                    .where(or_(
                        func.json_extract(Payment.payment_metadata, '$.flow') == 'food_scanner_credits',
                        func.json_extract(Payment.payment_metadata, '$.flow') == 'food_scanner_credits_razorpay'
                    ))
                    .where(~Client.contact.in_(EXCLUDED_CONTACTS))
                )
                if start_date:
                    ai_total_stmt = ai_total_stmt.where(func.date(Payment.captured_at) >= start_date_obj)
                if end_date:
                    ai_total_stmt = ai_total_stmt.where(func.date(Payment.captured_at) <= end_date_obj)
                if location_client_ids:
                    location_customer_ids = {str(cid) for cid in location_client_ids}
                    ai_total_stmt = ai_total_stmt.where(Payment.customer_id.in_(location_customer_ids))
                    
                ai_total_result = await db.execute(ai_total_stmt)
                ai_credits_total_purchases = ai_total_result.scalar() or 0

                # Build purchases over time (graph only)
                for row in ai_credits_results:
                    date_key = row.purchase_date.isoformat() if row.purchase_date else None
                    if date_key:
                        if date_key not in all_purchases_over_time:
                            all_purchases_over_time[date_key] = 0
                        all_purchases_over_time[date_key] += row.purchase_count
                        category_breakdown["ai_credits"]["purchases_over_time"].append({
                            "date": date_key,
                            "purchases": row.purchase_count
                        })

                category_breakdown["ai_credits"]["purchases"] = ai_credits_total_purchases
                category_breakdown["ai_credits"]["unique_users"] = ai_credits_unique_users_count

            except Exception as e:
                import logging
                logging.error(f"Purchase analytics - AI Credits error: {str(e)}")
                pass

        # 3.6. AI DIET COACH PURCHASES
        if (not source or source == "ai_diet_coach") and not gym_id:
            try:
                ai_diet_coach_stmt = (
                    select(
                        func.date(Payment.captured_at).label('purchase_date'),
                        func.count().label('purchase_count'),
                        func.count(distinct(Payment.customer_id)).label('unique_users')
                    )
                    .select_from(Payment)
                    .outerjoin(Client, Payment.customer_id == Client.client_id)
                    .where(Payment.status == "captured")
                    .where(func.json_extract(Payment.payment_metadata, '$.flow') == 'ai_diet_coach')
                    .where(func.date(Payment.captured_at) >= start_date_obj)
                    .where(func.date(Payment.captured_at) <= end_date_obj)
                    .where(~Client.contact.in_(EXCLUDED_CONTACTS))
                )

                if location_client_ids:
                    location_customer_ids = {str(cid) for cid in location_client_ids}
                    ai_diet_coach_stmt = ai_diet_coach_stmt.where(Payment.customer_id.in_(location_customer_ids))

                ai_diet_coach_stmt = ai_diet_coach_stmt.group_by(func.date(Payment.captured_at))

                result = await db.execute(ai_diet_coach_stmt)
                ai_diet_coach_results = result.all()

                # Get total unique users across all dates
                ai_diet_coach_unique_users_stmt = (
                    select(func.count(distinct(Payment.customer_id)))
                    .select_from(Payment)
                    .outerjoin(Client, Payment.customer_id == Client.client_id)
                    .where(Payment.status == "captured")
                    .where(func.json_extract(Payment.payment_metadata, '$.flow') == 'ai_diet_coach')
                    .where(func.date(Payment.captured_at) >= start_date_obj)
                    .where(func.date(Payment.captured_at) <= end_date_obj)
                    .where(~Client.contact.in_(EXCLUDED_CONTACTS))
                )

                if location_client_ids:
                    location_customer_ids = {str(cid) for cid in location_client_ids}
                    ai_diet_coach_unique_users_stmt = ai_diet_coach_unique_users_stmt.where(Payment.customer_id.in_(location_customer_ids))

                ai_diet_coach_unique_result = await db.execute(ai_diet_coach_unique_users_stmt)
                ai_diet_coach_unique_users_count = ai_diet_coach_unique_result.scalar() or 0

                ai_diet_total_stmt = (
                    select(func.count(Payment.id))
                    .select_from(Payment)
                    .outerjoin(Client, Payment.customer_id == Client.client_id)
                    .where(Payment.status == "captured")
                    .where(func.json_extract(Payment.payment_metadata, '$.flow') == 'ai_diet_coach')
                    .where(~Client.contact.in_(EXCLUDED_CONTACTS))
                )
                if start_date:
                    ai_diet_total_stmt = ai_diet_total_stmt.where(func.date(Payment.captured_at) >= start_date_obj)
                if end_date:
                    ai_diet_total_stmt = ai_diet_total_stmt.where(func.date(Payment.captured_at) <= end_date_obj)
                if location_client_ids:
                    location_customer_ids = {str(cid) for cid in location_client_ids}
                    ai_diet_total_stmt = ai_diet_total_stmt.where(Payment.customer_id.in_(location_customer_ids))
                    
                ai_diet_total_result = await db.execute(ai_diet_total_stmt)
                ai_diet_coach_total_purchases = ai_diet_total_result.scalar() or 0

                for row in ai_diet_coach_results:
                    date_key = row.purchase_date.isoformat() if row.purchase_date else None
                    if date_key:
                        if date_key not in all_purchases_over_time:
                            all_purchases_over_time[date_key] = 0
                        all_purchases_over_time[date_key] += row.purchase_count
                        category_breakdown["ai_diet_coach"]["purchases_over_time"].append({
                            "date": date_key,
                            "purchases": row.purchase_count
                        })

                category_breakdown["ai_diet_coach"]["purchases"] = ai_diet_coach_total_purchases
                category_breakdown["ai_diet_coach"]["unique_users"] = ai_diet_coach_unique_users_count

            except Exception as e:
                import logging
                logging.error(f"Purchase analytics - AI Diet Coach error: {str(e)}")
                pass

        # 4. GYM MEMBERSHIP PURCHASES
        # GMV fix: SQL EXISTS subquery with Gym JOIN + Client JOIN (replaces Python-loop in-memory filtering)
        if not source or source == "gym_membership":
            try:
                gym_meta_cond = or_(
                    func.json_unquote(func.json_extract(Order.order_metadata, "$.audit.source")) == "dailypass_checkout_api",
                    func.json_unquote(func.json_extract(Order.order_metadata, "$.order_info.flow")) == "unified_gym_membership_with_sub",
                    func.json_unquote(func.json_extract(Order.order_metadata, "$.order_info.flow")) == "unified_gym_membership_with_free_fittbot",
                    func.json_unquote(func.json_extract(Order.order_metadata, "$.order_info.flow")) == "gym_membership_with_bonus_credits",
                    func.json_unquote(func.json_extract(Order.order_metadata, "$.order_info.flow")) == "personal_training_with_bonus_credits"
                )

                # EXISTS on OrderItem + Gym JOIN — confirms gym physically exists
                gym_exists_cond = (
                    select(1)
                    .select_from(OrderItem)
                    .join(Gym, Gym.gym_id == cast(OrderItem.gym_id, Integer))
                    .where(
                        OrderItem.order_id == Order.id,
                        OrderItem.gym_id.isnot(None),
                        OrderItem.gym_id != "",
                        OrderItem.gym_id != "1"
                    )
                    .exists()
                )

                gm_base_conditions = [
                    Payment.status == "captured",
                    Order.status == "paid",
                    Order.customer_id.isnot(None),
                    gym_meta_cond,
                    gym_exists_cond,
                    func.date(Payment.captured_at) >= start_date_obj,
                    func.date(Payment.captured_at) <= end_date_obj,
                ]

                if gym_id:
                    gm_base_conditions.append(
                        select(1).select_from(OrderItem)
                        .where(OrderItem.order_id == Order.id, OrderItem.gym_id == str(gym_id))
                        .exists()
                    )

                if location_client_ids:
                    location_customer_ids = {str(cid) for cid in location_client_ids}
                    gm_base_conditions.append(Order.customer_id.in_(location_customer_ids))

                # Deduped subquery for counts
                gm_subq = (
                    select(
                        Order.id.label("order_id"),
                        Payment.captured_at.label("captured_at"),
                        Order.customer_id.label("customer_id")
                    )
                    .select_from(Payment)
                    .join(Order, Order.id == Payment.order_id)
                    .join(Client, Client.client_id == cast(Order.customer_id, Integer))
                    .where(*gm_base_conditions)
                    .distinct()
                    .subquery()
                )

                gm_total_stmt = select(
                    func.count(gm_subq.c.order_id).label("purchases"),
                    func.count(distinct(gm_subq.c.customer_id)).label("unique_users")
                ).select_from(gm_subq)

                gm_total_result = await db.execute(gm_total_stmt)
                gm_total_row = gm_total_result.one()
                category_breakdown["gym_membership"]["purchases"] = gm_total_row.purchases or 0
                category_breakdown["gym_membership"]["unique_users"] = gm_total_row.unique_users or 0

                # Time-series
                gm_time_stmt = (
                    select(
                        func.date(gm_subq.c.captured_at).label("purchase_date"),
                        func.count().label("purchase_count")
                    )
                    .select_from(gm_subq)
                    .group_by(func.date(gm_subq.c.captured_at))
                )
                gm_time_result = await db.execute(gm_time_stmt)
                for row in gm_time_result.all():
                    date_key = row.purchase_date.isoformat() if row.purchase_date else None
                    if date_key:
                        if date_key not in all_purchases_over_time:
                            all_purchases_over_time[date_key] = 0
                        all_purchases_over_time[date_key] += row.purchase_count
                        category_breakdown["gym_membership"]["purchases_over_time"].append({
                            "date": date_key,
                            "purchases": row.purchase_count
                        })
                category_breakdown["gym_membership"]["purchases_over_time"].sort(key=lambda x: x["date"])

                # Location tracking — re-added: get Client.location for each gym_membership order
                if not location or location == "all":
                    try:
                        gm_customer_ids_stmt = select(gm_subq.c.customer_id).select_from(gm_subq)
                        gm_customer_ids_result = await db.execute(gm_customer_ids_stmt)
                        gm_customer_ids = list(set([str(row[0]) for row in gm_customer_ids_result.all() if row[0]]))

                        if gm_customer_ids:
                            gm_loc_stmt = select(Client.client_id, Client.location).where(
                                Client.client_id.in_(gm_customer_ids),
                                Client.location.isnot(None),
                                Client.location != ''
                            )
                            gm_loc_result = await db.execute(gm_loc_stmt)
                            gm_client_locations = {str(row[0]): row[1] for row in gm_loc_result.all()}

                            for cid in gm_customer_ids:
                                if cid in gm_client_locations:
                                    normalized_loc = gm_client_locations[cid].strip().replace(' ', '_')
                                    if normalized_loc not in location_purchases:
                                        location_purchases[normalized_loc] = 0
                                    location_purchases[normalized_loc] += 1
                        
                        # GMV fix: Also track gym-wise purchases for Gym Memberships
                        gm_gym_stmt = (
                            select(
                                OrderItem.gym_id,
                                func.count().label('purchase_count')
                            )
                            .select_from(Payment)
                            .join(Order, Order.id == Payment.order_id)
                            .join(OrderItem, OrderItem.order_id == Order.id)
                            .join(Gym, Gym.gym_id == cast(OrderItem.gym_id, Integer))
                            .where(
                                Payment.status == "captured",
                                Order.status == "paid",
                                OrderItem.gym_id.isnot(None),
                                OrderItem.gym_id != "1",
                                func.date(Payment.captured_at) >= start_date_obj,
                                func.date(Payment.captured_at) <= end_date_obj,
                                gym_meta_cond
                            )
                        )
                        
                        if location_client_ids:
                            gm_gym_stmt = gm_gym_stmt.where(Order.customer_id.in_(location_client_ids))
                            
                        gm_gym_stmt = gm_gym_stmt.group_by(OrderItem.gym_id)
                        gm_gym_result = await db.execute(gm_gym_stmt)
                        for row in gm_gym_result.all():
                            try:
                                g_id = int(row.gym_id)
                                if g_id not in gym_purchases:
                                    gym_purchases[g_id] = 0
                                gym_purchases[g_id] += row.purchase_count
                            except (ValueError, TypeError):
                                pass

                    except Exception as e:
                        import logging
                        logging.error(f"Error getting location-wise gym membership purchases: {str(e)}")

            except Exception as e:
                import logging
                logging.error(f"Purchase analytics - Gym Membership error: {str(e)}")
                pass

        # Convert all purchases over time to sorted array
        purchases_over_time = [
            {
                "date": date,
                "purchases": count
            }
            for date, count in sorted(all_purchases_over_time.items())
        ]

        # Calculate total purchases across all categories
        total_purchases = sum(cat_data["purchases"] for cat_data in category_breakdown.values())

        # Debug logging
        import logging
        logging.info(f"Purchase analytics total_purchases: {total_purchases}, gym_purchases: {gym_purchases}")

        # Build gym breakdown
        gym_breakdown = []
        if not gym_id and gym_purchases:
            # Build full gym breakdown when no gym filter is applied
            gym_names = {}
            gym_ids = list(gym_purchases.keys())
            gym_stmt = select(Gym.gym_id, Gym.name).where(Gym.gym_id.in_(gym_ids))
            gym_result = await db.execute(gym_stmt)
            for gym_id_val, gym_name in gym_result.all():
                gym_names[gym_id_val] = gym_name

            gym_breakdown = [
                {
                    "gym_id": gym_id,
                    "gym_name": gym_names.get(gym_id, f"Gym {gym_id}"),
                    "revenue": gym_purchases[gym_id]
                }
                for gym_id in sorted(gym_purchases.keys(), key=lambda x: gym_purchases[x], reverse=True)
            ]
        elif gym_id:
            # When gym filter is applied, include the filtered gym in breakdown
            gym_names = {}
            gym_stmt = select(Gym.gym_id, Gym.name).where(Gym.gym_id == gym_id)
            gym_result = await db.execute(gym_stmt)
            for gym_id_val, gym_name in gym_result.all():
                gym_names[gym_id_val] = gym_name

            # Use the gym_purchases count if available, otherwise 0
            gym_purchases_count = gym_purchases.get(gym_id, 0)

            gym_breakdown = [{
                "gym_id": gym_id,
                "gym_name": gym_names.get(gym_id, f"Gym {gym_id}"),
                "revenue": gym_purchases_count
            }]
            logging.info(f"Gym breakdown for filtered gym {gym_id}: {gym_breakdown}")

        # Build location breakdown
        location_breakdown = []
        if location_purchases:
            # Sort locations by purchase count (descending)
            location_breakdown = [
                {
                    "location": loc,
                    "purchases": count
                }
                for loc, count in sorted(location_purchases.items(), key=lambda x: x[1], reverse=True)
            ]
            logging.info(f"Location breakdown: {location_breakdown}")

        # Build revenue by city breakdown — uses same GMV source logic as compute_gmv_totals()
        # Sums revenue from all 5 sources grouped by Gym.city
        revenue_by_city = []
        try:
            EXCLUDED_CONTACTS_SET = ["7373675762", "9486987082", "8667458723", "9840633149", "8667427956", "8667488723", "7975847236"]
            city_revenue_map = {}  # city -> total_revenue (rupees)

            # ── 1. Daily Pass ────────────────────────────────────────────────────
            try:
                _dp_session = get_dailypass_session()
                dp_city_q = (
                    _dp_session.query(
                        Gym.city.label("city"),
                        func.coalesce(func.sum(Payment.amount_minor / 100.0), 0).label("revenue")
                    )
                    .join(Gym, Gym.gym_id == func.cast(DailyPass.gym_id, Integer), isouter=False)
                    .outerjoin(Payment, DailyPass.payment_id == Payment.provider_payment_id)
                    .filter(
                        DailyPass.gym_id != "1",
                        func.date(DailyPass.created_at) >= start_date_obj,
                        func.date(DailyPass.created_at) <= end_date_obj,
                    )
                    .group_by(Gym.city)
                )
                if gym_id:
                    dp_city_q = dp_city_q.filter(DailyPass.gym_id == str(gym_id))
                for row in dp_city_q.all():
                    city = (row.city or "Unknown").strip()
                    city_revenue_map[city] = city_revenue_map.get(city, 0) + float(row.revenue or 0)
                _dp_session.close()
            except Exception as e:
                logging.error(f"[RevCity] daily_pass error: {e}")

            # ── 2. Sessions (SessionPurchase) ────────────────────────────────────
            try:
                sess_city_stmt = (
                    select(
                        func.coalesce(Gym.city, "Unknown").label("city"),
                        func.coalesce(func.sum(SessionPurchase.payable_rupees), 0).label("revenue")
                    )
                    .select_from(SessionPurchase)
                    .join(Gym, SessionPurchase.gym_id == Gym.gym_id)
                    .where(
                        SessionPurchase.status == "paid",
                        SessionPurchase.gym_id != 1,
                        func.date(SessionPurchase.created_at) >= start_date_obj,
                        func.date(SessionPurchase.created_at) <= end_date_obj,
                    )
                    .group_by(Gym.city)
                )
                if gym_id:
                    sess_city_stmt = sess_city_stmt.where(SessionPurchase.gym_id == gym_id)
                sess_city_result = await db.execute(sess_city_stmt)
                for row in sess_city_result.all():
                    city = (row.city or "Unknown").strip()
                    city_revenue_map[city] = city_revenue_map.get(city, 0) + float(row.revenue or 0)
            except Exception as e:
                logging.error(f"[RevCity] sessions error: {e}")

            # ── 3. Nutrition Plans ───────────────────────────────────────────────
            # No city/gym dimension — assign to "App" bucket
            try:
                nutri_city_stmt = (
                    select(func.coalesce(func.sum(Payment.amount_minor / 100.0), 0).label("revenue"))
                    .select_from(Payment)
                    .outerjoin(Client, Payment.customer_id == Client.client_id)
                    .join(
                        NutritionEligibility,
                        cast(Payment.order_id, String) == NutritionEligibility.source_id
                    )
                    .where(
                        NutritionEligibility.source_type == "fymble_purchase",
                        Payment.status == "captured",
                        or_(
                            func.json_extract(Payment.payment_metadata, "$.flow") == "nutrition_purchase_googleplay",
                            func.json_extract(Payment.payment_metadata, "$.flow") == "nutrition_package_razorpay",
                            func.json_extract(Payment.payment_metadata, "$.flow") == "basic_nutrition_plan",
                            func.json_extract(Payment.payment_metadata, "$.flow") == "expert_nutrition_plan",
                            func.json_extract(Payment.payment_metadata, "$.flow") == "elite_nutrition_plan"
                        ),
                        func.date(Payment.captured_at) >= start_date_obj,
                        func.date(Payment.captured_at) <= end_date_obj,
                        ~Client.contact.in_(EXCLUDED_CONTACTS_SET),
                    )
                )
                nutri_result = await db.execute(nutri_city_stmt)
                nutri_rev = float(nutri_result.scalar() or 0)
                if nutri_rev > 0:
                    city_revenue_map["App"] = city_revenue_map.get("App", 0) + nutri_rev
            except Exception as e:
                logging.error(f"[RevCity] nutrition error: {e}")

            # ── 4. AI Credits ────────────────────────────────────────────────────
            # No city/gym dimension — assign to "App" bucket
            try:
                ai_city_stmt = (
                    select(func.coalesce(func.sum(Payment.amount_minor / 100.0), 0).label("revenue"))
                    .select_from(Payment)
                    .outerjoin(Client, Payment.customer_id == Client.client_id)
                    .where(
                        Payment.status == "captured",
                        or_(
                            func.json_extract(Payment.payment_metadata, "$.flow") == "food_scanner_credits",
                            func.json_extract(Payment.payment_metadata, "$.flow") == "food_scanner_credits_razorpay"
                        ),
                        func.date(Payment.captured_at) >= start_date_obj,
                        func.date(Payment.captured_at) <= end_date_obj,
                        ~Client.contact.in_(EXCLUDED_CONTACTS_SET),
                    )
                )
                ai_result = await db.execute(ai_city_stmt)
                ai_rev = float(ai_result.scalar() or 0)
                if ai_rev > 0:
                    city_revenue_map["App"] = city_revenue_map.get("App", 0) + ai_rev
            except Exception as e:
                logging.error(f"[RevCity] ai_credits error: {e}")

            # ── 5. Gym Membership ────────────────────────────────────────────────
            try:
                gm_meta_cond = or_(
                    func.json_unquote(func.json_extract(Order.order_metadata, "$.audit.source")) == "dailypass_checkout_api",
                    func.json_unquote(func.json_extract(Order.order_metadata, "$.order_info.flow")) == "unified_gym_membership_with_sub",
                    func.json_unquote(func.json_extract(Order.order_metadata, "$.order_info.flow")) == "unified_gym_membership_with_free_fittbot",
                )
                gm_exists = (
                    select(1)
                    .select_from(OrderItem)
                    .join(Gym, Gym.gym_id == cast(OrderItem.gym_id, Integer))
                    .where(
                        OrderItem.order_id == Order.id,
                        OrderItem.gym_id.isnot(None),
                        OrderItem.gym_id != "",
                        OrderItem.gym_id != "1",
                    )
                    .exists()
                )
                gm_city_conditions = [
                    Payment.status == "captured",
                    Order.status == "paid",
                    Order.customer_id.isnot(None),
                    gm_meta_cond,
                    gm_exists,
                    func.date(Payment.captured_at) >= start_date_obj,
                    func.date(Payment.captured_at) <= end_date_obj,
                ]
                if gym_id:
                    gm_city_conditions.append(
                        select(1).select_from(OrderItem)
                        .where(OrderItem.order_id == Order.id, OrderItem.gym_id == str(gym_id))
                        .exists()
                    )

                # Join through OrderItem to get the gym, then Gym.city
                gm_city_subq = (
                    select(
                        Order.id.label("order_id"),
                        Order.gross_amount_minor.label("gross_amount_minor"),
                        OrderItem.gym_id.label("item_gym_id"),
                    )
                    .select_from(Payment)
                    .join(Order, Order.id == Payment.order_id)
                    .join(Client, Client.client_id == cast(Order.customer_id, Integer))
                    .join(OrderItem, and_(
                        OrderItem.order_id == Order.id,
                        OrderItem.gym_id.isnot(None),
                        OrderItem.gym_id != "",
                        OrderItem.gym_id != "1",
                    ))
                    .where(*gm_city_conditions)
                    .distinct()
                    .subquery()
                )

                gm_city_stmt = (
                    select(
                        func.coalesce(Gym.city, "Unknown").label("city"),
                        func.coalesce(func.sum(gm_city_subq.c.gross_amount_minor / 100.0), 0).label("revenue"),
                    )
                    .select_from(gm_city_subq)
                    .join(Gym, Gym.gym_id == cast(gm_city_subq.c.item_gym_id, Integer))
                    .group_by(Gym.city)
                )
                gm_city_result = await db.execute(gm_city_stmt)
                for row in gm_city_result.all():
                    city = (row.city or "Unknown").strip()
                    city_revenue_map[city] = city_revenue_map.get(city, 0) + float(row.revenue or 0)
            except Exception as e:
                logging.error(f"[RevCity] gym_membership error: {e}")

            # Build sorted output
            revenue_by_city = sorted(
                [{"city": city, "amount": round(amt, 2)} for city, amt in city_revenue_map.items() if city],
                key=lambda x: x["amount"],
                reverse=True
            )[:20]

            logging.info(f"Final revenue_by_city: {revenue_by_city}")

        except Exception as e:
            logging.error(f"Error building revenue_by_city: {str(e)}")
            import traceback
            traceback.print_exc()


        # Build separate stable gym list for filter dropdown (gyms with purchases in current timeframe)
        all_available_gyms = []
        if gym_purchases:
            all_gym_ids = list(gym_purchases.keys())
            all_gym_stmt = select(Gym.gym_id, Gym.name).where(Gym.gym_id.in_(all_gym_ids))
            all_gym_res = await db.execute(all_gym_stmt)
            gym_names_map = {row.gym_id: row.name for row in all_gym_res.all()}
            
            all_available_gyms = [
                {
                    "gym_id": gid,
                    "gym_name": gym_names_map.get(gid, f"Gym {gid}")
                }
                for gid in sorted(gym_purchases.keys(), key=lambda x: gym_purchases[x], reverse=True)
            ]

        analytics_data = {
            "totalPurchases": total_purchases,
            "categoryBreakdown": category_breakdown,
            "purchasesOverTime": purchases_over_time,
            "gymBreakdown": gym_breakdown,
            "availableGyms": all_available_gyms,  # Providing stable list for filters
            "locationBreakdown": location_breakdown,
            "revenueByCity": revenue_by_city,
            "filters": {
                "startDate": start_date if start_date else "All Time",
                "endDate": end_date if end_date else "All Time",
                "source": source or "all",
                "gymId": gym_id or "all",
                "location": location or "all"
            }
        }

        return {
            "success": True,
            "data": analytics_data,
            "message": "Purchase analytics fetched successfully"
        }

    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"Invalid date format: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/booking-averages")
async def get_booking_averages(
    db: AsyncSession = Depends(get_async_db)
):
    try:
        today = datetime.now().date()

        # Helper function to get purchases by source for a date range using compute_actual_booking_counts
        async def get_purchases_by_source(start_date, end_date):
            """Get purchases from all sources for a date range, broken down by source using compute_actual_booking_counts."""
            from app.fittbot_admin_api.revenue_service import compute_actual_booking_counts
            totals = await compute_actual_booking_counts(db, start_date, end_date)
            return {
                "daily_pass": int(totals.get("daily_pass") or 0),
                "sessions": int(totals.get("session") or 0),
                "gym_membership": int(totals.get("gym_membership") or 0),
                "fittbot_subscription": int(totals.get("nutrition_plan") or 0),
                "ai_credits": int(totals.get("ai_credits") or 0),
                "ai_diet_coach": int(totals.get("ai_diet_coach") or 0)
            }

        # Helper to calculate average from list of source-wise data
        def calculate_source_averages(data_list):
            """Calculate average for each source across data points."""
            if not data_list:
                return {"daily_pass": 0, "sessions": 0, "gym_membership": 0, "fittbot_subscription": 0, "ai_credits": 0, "ai_diet_coach": 0}

            num_points = len(data_list)
            return {
                "daily_pass": round(sum(d["daily_pass"] for d in data_list) / num_points, 2),
                "sessions": round(sum(d["sessions"] for d in data_list) / num_points, 2),
                "gym_membership": round(sum(d["gym_membership"] for d in data_list) / num_points, 2),
                "fittbot_subscription": round(sum(d["fittbot_subscription"] for d in data_list) / num_points, 2),
                "ai_credits": round(sum(d["ai_credits"] for d in data_list) / num_points, 2),
                "ai_diet_coach": round(sum(d.get("ai_diet_coach", 0) for d in data_list) / num_points, 2)
            }

        # Calculate Daily Average (last 7 completed days, excluding today) with source breakdown
        daily_source_data = []
        for i in range(7):
            day_date = today - timedelta(days=i + 1)
            day_data = await get_purchases_by_source(day_date, day_date)
            daily_source_data.append(day_data)
        daily_average = calculate_source_averages(daily_source_data)

        # Calculate Weekly Average (last 4 completed calendar weeks: Monday to Sunday) with source breakdown
        weekly_source_data = []
        # Find the start of the current week (Monday)
        current_week_start = today - timedelta(days=today.weekday())
        for i in range(4):
            # i = 0: Previous calendar week (ends on preceding Sunday, starts on preceding Monday)
            # i = 1: 2 calendar weeks ago
            # i = 2: 3 calendar weeks ago
            # i = 3: 4 calendar weeks ago
            week_end = current_week_start - timedelta(days=(i * 7) + 1)
            week_start = week_end - timedelta(days=6)
            week_data = await get_purchases_by_source(week_start, week_end)
            weekly_source_data.append(week_data)
        weekly_average = calculate_source_averages(weekly_source_data)

        # Calculate Monthly Average (last 3 FULL months) with source breakdown
        monthly_source_data = []
        for i in range(3):
            # Last 3 FULL months - exclude current incomplete month
            # Month 1: Previous full month
            # Month 2: 2 months ago (full month)
            # Month 3: 3 months ago (full month)
            month_date = (today.replace(day=1) - timedelta(days=1)).replace(day=1)
            for _ in range(i):
                month_date = (month_date - timedelta(days=1)).replace(day=1)
            month_start = month_date
            # Get last day of the month
            next_month = month_date.replace(day=28) + timedelta(days=4)
            month_end = next_month - timedelta(days=next_month.day)

            month_data = await get_purchases_by_source(month_start, month_end)
            monthly_source_data.append(month_data)
        monthly_average = calculate_source_averages(monthly_source_data)

        # Calculate totals for display
        daily_total = round(sum(daily_average.values()), 2)
        weekly_total = round(sum(weekly_average.values()), 2)
        monthly_total = round(sum(monthly_average.values()), 2)

        return {
            "success": True,
            "data": {
                "dailyAverage": daily_total,
                "weeklyAverage": weekly_total,
                "monthlyAverage": monthly_total,
                "dailyBreakdown": daily_average,
                "weeklyBreakdown": weekly_average,
                "monthlyBreakdown": monthly_average
            },
            "message": "Booking averages fetched successfully"
        }

    except Exception as e:
        import logging
        logging.error(f"Error in booking-averages: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e)) 


@router.get("/price-change-notifications/pending")
async def get_pending_price_notifications(
    current_admin: Admins = Depends(get_current_admin_from_cookie),
    db: AsyncSession = Depends(get_async_db)
):
    try:
        from app.models.adminmodels import PriceChangeNotification
        from app.models.fittbot_models.gym import GymPriceChanges
        from app.models.fittbot_models import Gym
        from sqlalchemy import select, and_, exists, func
        from sqlalchemy.orm import aliased
        
        # Step 1: Sync Phase
        # Find all GymPriceChanges rows that do not have any notification created for their price_id
        gpc_alias = aliased(GymPriceChanges)
        subq = (
            select(gpc_alias.price_id)
            .join(PriceChangeNotification, PriceChangeNotification.price_change_id == gpc_alias.id)
            .scalar_subquery()
        )
        sync_stmt = (
            select(GymPriceChanges)
            .where(GymPriceChanges.price_id.not_in(subq))
            .order_by(GymPriceChanges.created_at.desc())
        )
        
        sync_result = await db.execute(sync_stmt)
        missing_changes = sync_result.scalars().all()
        
        seen_price_ids = set()
        new_notifications = []
        for gpc in missing_changes:
            if gpc.price_id not in seen_price_ids:
                seen_price_ids.add(gpc.price_id)
                new_notif = PriceChangeNotification(
                    price_change_id=gpc.id,
                    admin_id=None,
                    viewed_at=None
                )
                db.add(new_notif)
                new_notifications.append(new_notif)
                
        if new_notifications:
            await db.commit()
            
        # Step 2: Fetch Phase
        # Select GymPriceChanges records that have not been viewed by current admin
        min_notif_subq = (
            select(func.min(PriceChangeNotification.id))
            .where(PriceChangeNotification.price_change_id == GymPriceChanges.id)
            .correlate(GymPriceChanges)
            .scalar_subquery()
        )
        
        stmt = (
            select(
                PriceChangeNotification.id.label("notification_id"),
                PriceChangeNotification.price_change_id,
                PriceChangeNotification.created_at.label("notification_created_at"),
                GymPriceChanges.last_price,
                GymPriceChanges.latest_price,
                GymPriceChanges.type,
                GymPriceChanges.price_id,
                GymPriceChanges.created_at.label("price_changed_at"),
                Gym.name.label("gym_name"),
                Gym.gym_id
            )
            .join(GymPriceChanges, PriceChangeNotification.price_change_id == GymPriceChanges.id)
            .join(Gym, GymPriceChanges.gym_id == Gym.gym_id)
            .where(
                and_(
                    PriceChangeNotification.id == min_notif_subq,
                    ~exists(
                        select(PriceChangeNotification.id)
                        .where(
                            and_(
                                PriceChangeNotification.price_change_id == GymPriceChanges.id,
                                PriceChangeNotification.admin_id == current_admin.admin_id,
                                PriceChangeNotification.viewed_at.is_not(None)
                            )
                        )
                        .correlate(GymPriceChanges)
                    )
                )
            )
            .order_by(PriceChangeNotification.created_at.desc())
        )
        
        result = await db.execute(stmt)
        rows = result.all()
        
        notifications_data = []
        for row in rows:
            notifications_data.append({
                "notification_id": row.notification_id,
                "price_change_id": row.price_change_id,
                "gym_id": row.gym_id,
                "gym_name": row.gym_name,
                "last_price": row.last_price,
                "latest_price": row.latest_price,
                "type": row.type,
                "price_id": row.price_id,
                "price_changed_at": row.price_changed_at.isoformat() if row.price_changed_at else None,
                "created_at": row.notification_created_at.isoformat() if row.notification_created_at else None
            })
            
        return {
            "success": True,
            "count": len(notifications_data),
            "notifications": notifications_data,
            "message": "Pending price change notifications fetched successfully"
        }
    except Exception as e:
        import logging
        logging.error(f"Error fetching pending price notifications: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/price-change-notifications/mark-viewed")
async def mark_price_notifications_viewed(
    current_admin: Admins = Depends(get_current_admin_from_cookie),
    db: AsyncSession = Depends(get_async_db)
):
    try:
        from app.models.adminmodels import PriceChangeNotification
        from app.models.fittbot_models.gym import GymPriceChanges
        from sqlalchemy import select, and_, exists
        from datetime import datetime
        
        # 1. Fetch all pending price change IDs for the current admin
        # These are GymPriceChanges.id values that have NO viewed PriceChangeNotification for this admin.
        
        viewed_exists = exists().where(
            and_(
                PriceChangeNotification.price_change_id == GymPriceChanges.id,
                PriceChangeNotification.admin_id == current_admin.admin_id,
                PriceChangeNotification.viewed_at.is_not(None)
            )
        )
        
        notif_exists = exists().where(
            PriceChangeNotification.price_change_id == GymPriceChanges.id
        )
        
        pending_stmt = (
            select(GymPriceChanges.id)
            .where(
                and_(
                    notif_exists,
                    ~viewed_exists
                )
            )
        )
        
        pending_result = await db.execute(pending_stmt)
        pending_ids = pending_result.scalars().all()
        
        if pending_ids:
            for gpc_id in pending_ids:
                # Check if a notification already exists for this admin and this price change (even if viewed_at is null)
                admin_notif_stmt = (
                    select(PriceChangeNotification)
                    .where(
                        and_(
                            PriceChangeNotification.price_change_id == gpc_id,
                            PriceChangeNotification.admin_id == current_admin.admin_id
                        )
                    )
                )
                admin_notif_result = await db.execute(admin_notif_stmt)
                admin_notif = admin_notif_result.scalar_one_or_none()
                
                if admin_notif:
                    admin_notif.viewed_at = datetime.now()
                else:
                    # Check if there is an unassigned notification (admin_id is NULL) for this price change
                    unassigned_notif_stmt = (
                        select(PriceChangeNotification)
                        .where(
                            and_(
                                PriceChangeNotification.price_change_id == gpc_id,
                                PriceChangeNotification.admin_id.is_(None)
                            )
                        )
                    )
                    unassigned_notif_result = await db.execute(unassigned_notif_stmt)
                    unassigned_notif = unassigned_notif_result.scalar_one_or_none()
                    
                    if unassigned_notif:
                        unassigned_notif.admin_id = current_admin.admin_id
                        unassigned_notif.viewed_at = datetime.now()
                    else:
                        # Create a new viewed notification for this admin
                        new_notif = PriceChangeNotification(
                            price_change_id=gpc_id,
                            admin_id=current_admin.admin_id,
                            viewed_at=datetime.now()
                        )
                        db.add(new_notif)
            
            await db.commit()
        
        return {
            "success": True,
            "message": "All pending price notifications marked as viewed successfully"
        }
    except Exception as e:
        import logging
        logging.error(f"Error marking price notifications as viewed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# @router.get("/purchase-analytics")
# async def get_purchase_analytics(
#     start_date: str = None,
#     end_date: str = None,
#     source: str = None,
#     gym_id: int = None,
#     location: str = None,
#     db: AsyncSession = Depends(get_async_db)
# ):

#     try:
#         # Debug: Log the gym_id parameter
#         import logging
#         logging.info(f"Purchase analytics called with gym_id: {gym_id} (type: {type(gym_id)}), start_date: {start_date}, end_date: {end_date}, source: {source}")

#         # Parse dates if provided
#         if start_date:
#             start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
#         else:
#             # Default to early date for overall data
#             start_date_obj = datetime(1970, 1, 1).date()

#         if end_date:
#             end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
#         else:
#             # Default to today
#             end_date_obj = datetime(2100, 1, 1).date()

#         # Get client IDs filtered by location if location is provided
#         location_client_ids = set()
#         if location and location != "all":
#             try:
#                 location_stmt = select(Client.client_id).where(Client.location == location)
#                 location_result = await db.execute(location_stmt)
#                 location_client_ids = set(row[0] for row in location_result.all())
#             except Exception as e:
#                 location_client_ids = set()
#         else:
#             # Debug: Check how many purchases have NULL or invalid client_id
#             try:
#                 dailypass_session = get_dailypass_session()

#                 # Count total DailyPass purchases
#                 total_dp = dailypass_session.query(func.count()).scalar()

#                 # Count DailyPass with NULL client_id
#                 null_client_dp = dailypass_session.query(func.count()).filter(DailyPass.client_id.is_(None)).scalar()

#                 # Count DailyPass with client_id NOT in Client table
#                 valid_client_ids = select(Client.client_id)
#                 not_in_client = dailypass_session.query(func.count()).filter(~DailyPass.client_id.in_(valid_client_ids)).scalar()

#                 dailypass_session.close()
#             except Exception as e:
#                 pass

#         # Initialize result structure
#         category_breakdown = {
#             "daily_pass": {"purchases": 0, "unique_users": 0, "purchases_over_time": []},
#             "sessions": {"purchases": 0, "unique_users": 0, "purchases_over_time": []},
#             "fittbot_subscription": {"purchases": 0, "unique_users": 0, "purchases_over_time": []},
#             "ai_credits": {"purchases": 0, "unique_users": 0, "purchases_over_time": []},
#             "ai_diet_coach": {"purchases": 0, "unique_users": 0, "purchases_over_time": []},
#             "gym_membership": {"purchases": 0, "unique_users": 0, "purchases_over_time": []}
#         }

#         all_purchases_over_time = {}  # date -> total count
#         gym_purchases = {}  # gym_id -> purchase count
#         location_purchases = {}  # location -> purchase count

#         # 1. DAILY PASS PURCHASES - Single aggregated query
#         # GMV fix: INNER JOIN with Gym table to exclude orphaned DailyPass records (gym_id not in Gym table)
#         if not source or source == "daily_pass":
#             try:
#                 import logging
#                 logging.info(f"Daily pass purchases: gym_id={gym_id}, source={source}")
#                 dailypass_session = get_dailypass_session()

#                 # Build base query — JOIN Gym table to exclude records with no matching gym
#                 # Option A: SUM(days_total * head_count) so multipass purchases scale by headcount
#                 all_query = dailypass_session.query(
#                     func.coalesce(func.sum(DailyPass.days_total * DailyPass.head_count), 0).label('total_purchases'),
#                     func.count(distinct(DailyPass.client_id)).label('total_unique_users')
#                 ).join(
#                     Gym, Gym.gym_id == func.cast(DailyPass.gym_id, Integer), isouter=False
#                 ).filter(
#                     func.date(DailyPass.created_at) >= start_date_obj,
#                     func.date(DailyPass.created_at) <= end_date_obj,
#                     DailyPass.gym_id != "1"
#                 )

#                 # Apply gym filter if provided
#                 if gym_id:
#                     gym_id_str = str(gym_id)
#                     all_query = all_query.filter(DailyPass.gym_id == gym_id_str)
#                     logging.info(f"Daily pass query with gym filter: gym_id_str={gym_id_str}")

#                 # Apply location filter if provided
#                 if location_client_ids:
#                     all_query = all_query.filter(DailyPass.client_id.in_(location_client_ids))

#                 # Get totals
#                 total_result = all_query.first()

#                 if total_result:
#                     category_breakdown["daily_pass"]["purchases"] = total_result.total_purchases or 0
#                     category_breakdown["daily_pass"]["unique_users"] = total_result.total_unique_users or 0

#                 # Build separate query for purchases over time (grouped by date) — same Gym JOIN
#                 # Option A: SUM(days_total * head_count) so multipass purchases scale by headcount in time-series too
#                 base_query = dailypass_session.query(
#                     func.date(DailyPass.created_at).label('purchase_date'),
#                     func.coalesce(func.sum(DailyPass.days_total * DailyPass.head_count), 0).label('purchase_count')
#                 ).join(
#                     Gym, Gym.gym_id == func.cast(DailyPass.gym_id, Integer), isouter=False
#                 ).filter(
#                     func.date(DailyPass.created_at) >= start_date_obj,
#                     func.date(DailyPass.created_at) <= end_date_obj,
#                     DailyPass.gym_id != "1"
#                 )

#                 # Apply gym filter if provided
#                 if gym_id:
#                     base_query = base_query.filter(DailyPass.gym_id == str(gym_id))

#                 # Apply location filter if provided
#                 if location_client_ids:
#                     base_query = base_query.filter(DailyPass.client_id.in_(location_client_ids))

#                 # Group by date and get results
#                 base_query = base_query.group_by(func.date(DailyPass.created_at))
#                 time_result = base_query.all()

#                 if time_result:
#                     # Build purchases over time
#                     for row in time_result:
#                         date_key = row.purchase_date.isoformat() if row.purchase_date else None
#                         if date_key:
#                             if date_key not in all_purchases_over_time:
#                                 all_purchases_over_time[date_key] = 0
#                             all_purchases_over_time[date_key] += row.purchase_count
#                             category_breakdown["daily_pass"]["purchases_over_time"].append({
#                                 "date": date_key,
#                                 "purchases": row.purchase_count
#                             })

#                 # Sort purchases over time by date
#                 category_breakdown["daily_pass"]["purchases_over_time"].sort(key=lambda x: x["date"])

#                 # Get gym-wise purchases - ALWAYS get all gyms with purchases in timeframe/location
#                 # This ensures the filter dropdown remains populated even when a gym is selected.
#                 # Option A: SUM(days_total * head_count) so multipass purchases scale by headcount gym-wise
#                 gym_query = dailypass_session.query(
#                     DailyPass.gym_id,
#                     func.coalesce(func.sum(DailyPass.days_total * DailyPass.head_count), 0).label('purchase_count')
#                 ).join(
#                     Gym, Gym.gym_id == func.cast(DailyPass.gym_id, Integer), isouter=False
#                 ).filter(
#                     func.date(DailyPass.created_at) >= start_date_obj,
#                     func.date(DailyPass.created_at) <= end_date_obj,
#                     DailyPass.gym_id.isnot(None),
#                     DailyPass.gym_id != "1"  # Exclude gym_id = 1
#                 )

#                 # Apply location filter if provided
#                 if location_client_ids:
#                     gym_query = gym_query.filter(DailyPass.client_id.in_(location_client_ids))

#                 gym_query = gym_query.group_by(DailyPass.gym_id)
#                 gym_result = gym_query.all()

#                 for row in gym_result:
#                     try:
#                         gym_key = int(row.gym_id)
#                         if gym_key not in gym_purchases:
#                             gym_purchases[gym_key] = 0
#                         gym_purchases[gym_key] += row.purchase_count
#                     except (ValueError, TypeError):
#                         pass
                
#                 if gym_id:
#                     logging.info(f"Daily pass for filtered gym {gym_id}: {gym_purchases.get(int(gym_id), 0)} purchases")

#                 logging.info(f"Daily pass gym_purchases: {gym_purchases}")

#                 # Get location-wise purchases (only when no location filter is applied)
#                 # GMV fix: add Gym INNER JOIN so orphaned DailyPass records are excluded
#                 if not location or location == "all":
#                     try:
#                         # Get all DailyPass client_ids and their headcount-scaled purchases — with Gym JOIN to exclude orphans
#                         daily_pass_records = dailypass_session.query(
#                             DailyPass.client_id,
#                             (DailyPass.days_total * DailyPass.head_count).label("purchase_count")
#                         ).join(
#                             Gym, Gym.gym_id == func.cast(DailyPass.gym_id, Integer), isouter=False
#                         ).filter(
#                             func.date(DailyPass.created_at) >= start_date_obj,
#                             func.date(DailyPass.created_at) <= end_date_obj,
#                             DailyPass.client_id.isnot(None),
#                             DailyPass.gym_id != "1"
#                         )

#                         if gym_id:
#                             daily_pass_records = daily_pass_records.filter(DailyPass.gym_id == str(gym_id))

#                         daily_pass_records = daily_pass_records.all()

#                         client_ids = list(set([str(r.client_id) for r in daily_pass_records if r.client_id]))

#                         if client_ids:
#                             client_location_stmt = select(Client.client_id, Client.location).where(
#                                 Client.client_id.in_(client_ids),
#                                 Client.location.isnot(None),
#                                 Client.location != ''
#                             )
#                             client_location_result = await db.execute(client_location_stmt)
#                             client_locations = {str(row[0]): row[1] for row in client_location_result.all()}

#                             for record in daily_pass_records:
#                                 if record.client_id and str(record.client_id) in client_locations:
#                                     raw_loc = client_locations[str(record.client_id)]
#                                     normalized_loc = raw_loc.strip().replace(' ', '_')
#                                     if normalized_loc not in location_purchases:
#                                         location_purchases[normalized_loc] = 0
#                                     location_purchases[normalized_loc] += record.purchase_count

#                     except Exception as e:
#                         import logging
#                         logging.error(f"Error getting location-wise daily pass purchases: {str(e)}")


#                 dailypass_session.close()
#             except Exception as e:
#                 # Log error for debugging daily pass purchases
#                 import logging
#                 logging.error(f"Purchase analytics - Daily pass error: {str(e)}")
#                 pass

#         # 2. SESSION PURCHASES - Use SessionPurchase (status='paid') + Gym INNER JOIN (matches GMV logic)
#         if not source or source == "sessions":
#             try:
#                 # GMV fix: SessionPurchase with status='paid' + INNER JOIN Gym (no orphans)
#                 # SUM(sessions_count) so multi-class purchases count all classes (not just 1 per transaction)
#                 total_stmt = (
#                     select(
#                         func.coalesce(func.sum(SessionPurchase.sessions_count), 0).label('total_purchases'),
#                         func.count(distinct(SessionPurchase.client_id)).label('total_unique_users')
#                     )
#                     .select_from(SessionPurchase)
#                     .join(Gym, SessionPurchase.gym_id == Gym.gym_id)
#                     .where(
#                         SessionPurchase.status == "paid",
#                         SessionPurchase.gym_id != 1,
#                         func.date(SessionPurchase.created_at) >= start_date_obj,
#                         func.date(SessionPurchase.created_at) <= end_date_obj,
#                     )
#                 )

#                 # Apply gym filter if provided
#                 if gym_id:
#                     total_stmt = total_stmt.where(SessionPurchase.gym_id == gym_id)

#                 # Apply location filter if provided
#                 if location_client_ids:
#                     total_stmt = total_stmt.where(SessionPurchase.client_id.in_(location_client_ids))

#                 total_result = await db.execute(total_stmt)
#                 total_row = total_result.first()

#                 if total_row:
#                     category_breakdown["sessions"]["purchases"] = total_row.total_purchases or 0
#                     category_breakdown["sessions"]["unique_users"] = total_row.total_unique_users or 0

#                 # Then get purchases over time (grouped by date) — same model/filters
#                 # GMV fix: SUM(sessions_count) per day in time-series
#                 time_stmt = (
#                     select(
#                         func.date(SessionPurchase.created_at).label('purchase_date'),
#                         func.coalesce(func.sum(SessionPurchase.sessions_count), 0).label('purchase_count')
#                     )
#                     .select_from(SessionPurchase)
#                     .join(Gym, SessionPurchase.gym_id == Gym.gym_id)
#                     .where(
#                         SessionPurchase.status == "paid",
#                         SessionPurchase.gym_id != 1,
#                         func.date(SessionPurchase.created_at) >= start_date_obj,
#                         func.date(SessionPurchase.created_at) <= end_date_obj,
#                     )
#                 )

#                 # Apply gym filter if provided
#                 if gym_id:
#                     time_stmt = time_stmt.where(SessionPurchase.gym_id == gym_id)

#                 # Apply location filter if provided
#                 if location_client_ids:
#                     time_stmt = time_stmt.where(SessionPurchase.client_id.in_(location_client_ids))

#                 # Group by date and execute
#                 time_stmt = time_stmt.group_by(func.date(SessionPurchase.created_at))
#                 time_result = await db.execute(time_stmt)
#                 session_results = time_result.all()

#                 if session_results:
#                     # Build purchases over time
#                     for row in session_results:
#                         date_key = row.purchase_date.isoformat() if row.purchase_date else None
#                         if date_key:
#                             if date_key not in all_purchases_over_time:
#                                 all_purchases_over_time[date_key] = 0
#                             all_purchases_over_time[date_key] += row.purchase_count
#                             category_breakdown["sessions"]["purchases_over_time"].append({
#                                 "date": date_key,
#                                 "purchases": row.purchase_count
#                             })

#                 # Sort purchases over time by date
#                 category_breakdown["sessions"]["purchases_over_time"].sort(key=lambda x: x["date"])

#                 # Get gym-wise purchases — ALWAYS get all gyms (for filter list stability)
#                 gym_wise_stmt = (
#                     select(
#                         SessionPurchase.gym_id,
#                         func.count().label('purchase_count')
#                     )
#                     .select_from(SessionPurchase)
#                     .join(Gym, SessionPurchase.gym_id == Gym.gym_id)
#                     .where(
#                         SessionPurchase.status == "paid",
#                         func.date(SessionPurchase.created_at) >= start_date_obj,
#                         func.date(SessionPurchase.created_at) <= end_date_obj,
#                         SessionPurchase.gym_id.isnot(None),
#                         SessionPurchase.gym_id != 1
#                     )
#                 )

#                 if location_client_ids:
#                     gym_wise_stmt = gym_wise_stmt.where(SessionPurchase.client_id.in_(location_client_ids))

#                 gym_wise_stmt = gym_wise_stmt.group_by(SessionPurchase.gym_id)
#                 gym_wise_result = await db.execute(gym_wise_stmt)
#                 gym_wise_rows = gym_wise_result.all()

#                 for row in gym_wise_rows:
#                     if row.gym_id:
#                         if row.gym_id not in gym_purchases:
#                             gym_purchases[row.gym_id] = 0
#                         gym_purchases[row.gym_id] += row.purchase_count
                
#                 if gym_id:
#                     logging.info(f"Sessions for filtered gym {gym_id}: {gym_purchases.get(int(gym_id), 0)} purchases")

#                 # Get location-wise purchases (only when no location filter is applied)
#                 if not location or location == "all":
#                     try:
#                         session_records_stmt = (
#                             select(SessionPurchase.client_id)
#                             .select_from(SessionPurchase)
#                             .join(Gym, SessionPurchase.gym_id == Gym.gym_id)
#                             .where(
#                                 SessionPurchase.status == "paid",
#                                 func.date(SessionPurchase.created_at) >= start_date_obj,
#                                 func.date(SessionPurchase.created_at) <= end_date_obj,
#                                 SessionPurchase.client_id.isnot(None),
#                                 SessionPurchase.gym_id != 1
#                             )
#                         )

#                         if gym_id:
#                             session_records_stmt = session_records_stmt.where(SessionPurchase.gym_id == gym_id)

#                         if location_client_ids:
#                             session_records_stmt = session_records_stmt.where(SessionPurchase.client_id.in_(location_client_ids))

#                         session_records_result = await db.execute(session_records_stmt)
#                         session_records = session_records_result.all()

#                         # Get unique client_ids
#                         client_ids = list(set([str(row[0]) for row in session_records]))

#                         if client_ids:
#                             # Query Client table to get locations
#                             client_location_stmt = select(Client.client_id, Client.location).where(
#                                 Client.client_id.in_(client_ids),
#                                 Client.location.isnot(None),
#                                 Client.location != ''
#                             )
#                             client_location_result = await db.execute(client_location_stmt)
#                             client_locations = {str(row[0]): row[1] for row in client_location_result.all()}

#                             # Count purchases per location (normalize location names)
#                             for row in session_records:
#                                 if row[0] and str(row[0]) in client_locations:
#                                     raw_loc = client_locations[str(row[0])]
#                                     # Normalize location: trim whitespace and replace spaces with underscores
#                                     normalized_loc = raw_loc.strip().replace(' ', '_')
#                                     if normalized_loc not in location_purchases:
#                                         location_purchases[normalized_loc] = 0
#                                     location_purchases[normalized_loc] += 1

#                     except Exception as e:
#                         import logging
#                         logging.error(f"Error getting location-wise session purchases: {str(e)}")

#             except Exception:
#                 pass

#         # 3. NUTRITIONIST PLAN (FITTBOT SUBSCRIPTION) PURCHASES
#         # GMV fix: added excluded contacts filter to match gmv-summary logic
#         # NOTE: Skip when gym filter is applied (not gym-specific purchases)
#         EXCLUDED_CONTACTS = ["7373675762", "9486987082", "8667458723", "9840633149", "8667427956", "8667488723", "7975847236"]
#         if (not source or source == "fittbot_subscription") and not gym_id:
#             try:
#                 nutritionist_stmt = (
#                     select(
#                         func.date(Payment.captured_at).label('purchase_date'),
#                         func.count().label('purchase_count'),
#                         func.count(distinct(Payment.customer_id)).label('unique_users')
#                     )
#                     .select_from(Payment)
#                     .outerjoin(Client, Payment.customer_id == Client.client_id)
#                     .join(
#                         NutritionEligibility,
#                         cast(Payment.order_id, String) == NutritionEligibility.source_id
#                     )
#                     .where(NutritionEligibility.source_type == "fymble_purchase")
#                     .where(Payment.status == "captured")
#                     .where(or_(
#                         func.json_extract(Payment.payment_metadata, '$.flow') == 'nutrition_purchase_googleplay',
#                         func.json_extract(Payment.payment_metadata, '$.flow') == 'nutrition_package_razorpay',
#                         func.json_extract(Payment.payment_metadata, '$.flow') == 'basic_nutrition_plan',
#                         func.json_extract(Payment.payment_metadata, '$.flow') == 'expert_nutrition_plan',
#                         func.json_extract(Payment.payment_metadata, '$.flow') == 'elite_nutrition_plan'
#                     ))
#                     .where(func.date(Payment.captured_at) >= start_date_obj)
#                     .where(func.date(Payment.captured_at) <= end_date_obj)
#                     .where(~Client.contact.in_(EXCLUDED_CONTACTS))
#                 )

#                 if location_client_ids:
#                     location_customer_ids = {str(cid) for cid in location_client_ids}
#                     nutritionist_stmt = nutritionist_stmt.where(Payment.customer_id.in_(location_customer_ids))

#                 nutritionist_stmt = nutritionist_stmt.group_by(func.date(Payment.captured_at))

#                 result = await db.execute(nutritionist_stmt)
#                 nutritionist_results = result.all()

#                 # Get total unique users across all dates
#                 unique_users_stmt = (
#                     select(func.count(distinct(Payment.customer_id)))
#                     .select_from(Payment)
#                     .outerjoin(Client, Payment.customer_id == Client.client_id)
#                     .join(
#                         NutritionEligibility,
#                         cast(Payment.order_id, String) == NutritionEligibility.source_id
#                     )
#                     .where(NutritionEligibility.source_type == "fymble_purchase")
#                     .where(Payment.status == "captured")
#                     .where(or_(
#                         func.json_extract(Payment.payment_metadata, '$.flow') == 'nutrition_purchase_googleplay',
#                         func.json_extract(Payment.payment_metadata, '$.flow') == 'nutrition_package_razorpay',
#                         func.json_extract(Payment.payment_metadata, '$.flow') == 'basic_nutrition_plan',
#                         func.json_extract(Payment.payment_metadata, '$.flow') == 'expert_nutrition_plan',
#                         func.json_extract(Payment.payment_metadata, '$.flow') == 'elite_nutrition_plan'
#                     ))
#                     .where(func.date(Payment.captured_at) >= start_date_obj)
#                     .where(func.date(Payment.captured_at) <= end_date_obj)
#                     .where(~Client.contact.in_(EXCLUDED_CONTACTS))
#                 )

#                 if location_client_ids:
#                     location_customer_ids = {str(cid) for cid in location_client_ids}
#                     unique_users_stmt = unique_users_stmt.where(Payment.customer_id.in_(location_customer_ids))

#                 unique_result = await db.execute(unique_users_stmt)
#                 unique_users_count = unique_result.scalar() or 0

#                 # Calculate totals and build purchases over time
#                 total_purchases = 0
#                 for row in nutritionist_results:
#                     date_key = row.purchase_date.isoformat() if row.purchase_date else None
#                     if date_key:
#                         total_purchases += row.purchase_count
#                         if date_key not in all_purchases_over_time:
#                             all_purchases_over_time[date_key] = 0
#                         all_purchases_over_time[date_key] += row.purchase_count
#                         category_breakdown["fittbot_subscription"]["purchases_over_time"].append({
#                             "date": date_key,
#                             "purchases": row.purchase_count
#                         })

#                 category_breakdown["fittbot_subscription"]["purchases"] = total_purchases
#                 category_breakdown["fittbot_subscription"]["unique_users"] = unique_users_count

#             except Exception as e:
#                 import logging
#                 logging.error(f"Purchase analytics - Nutritionist Plan error: {str(e)}")
#                 pass

#         # 3.5. AI CREDITS PURCHASES
#         # GMV fix: added excluded contacts filter to match gmv-summary logic
#         # NOTE: Skip when gym filter is applied (not gym-specific purchases)
#         if (not source or source == "ai_credits") and not gym_id:
#             try:
#                 ai_credits_stmt = (
#                     select(
#                         func.date(Payment.captured_at).label('purchase_date'),
#                         func.count().label('purchase_count'),
#                         func.count(distinct(Payment.customer_id)).label('unique_users')
#                     )
#                     .select_from(Payment)
#                     .outerjoin(Client, Payment.customer_id == Client.client_id)
#                     .where(Payment.status == "captured")
#                     .where(or_(
#                         func.json_extract(Payment.payment_metadata, '$.flow') == 'food_scanner_credits',
#                         func.json_extract(Payment.payment_metadata, '$.flow') == 'food_scanner_credits_razorpay'
#                     ))
#                     .where(func.date(Payment.captured_at) >= start_date_obj)
#                     .where(func.date(Payment.captured_at) <= end_date_obj)
#                     .where(~Client.contact.in_(EXCLUDED_CONTACTS))
#                 )

#                 if location_client_ids:
#                     location_customer_ids = {str(cid) for cid in location_client_ids}
#                     ai_credits_stmt = ai_credits_stmt.where(Payment.customer_id.in_(location_customer_ids))

#                 ai_credits_stmt = ai_credits_stmt.group_by(func.date(Payment.captured_at))

#                 result = await db.execute(ai_credits_stmt)
#                 ai_credits_results = result.all()

#                 # Get total unique users across all dates
#                 ai_credits_unique_users_stmt = (
#                     select(func.count(distinct(Payment.customer_id)))
#                     .select_from(Payment)
#                     .outerjoin(Client, Payment.customer_id == Client.client_id)
#                     .where(Payment.status == "captured")
#                     .where(or_(
#                         func.json_extract(Payment.payment_metadata, '$.flow') == 'food_scanner_credits',
#                         func.json_extract(Payment.payment_metadata, '$.flow') == 'food_scanner_credits_razorpay'
#                     ))
#                     .where(func.date(Payment.captured_at) >= start_date_obj)
#                     .where(func.date(Payment.captured_at) <= end_date_obj)
#                     .where(~Client.contact.in_(EXCLUDED_CONTACTS))
#                 )

#                 if location_client_ids:
#                     location_customer_ids = {str(cid) for cid in location_client_ids}
#                     ai_credits_unique_users_stmt = ai_credits_unique_users_stmt.where(Payment.customer_id.in_(location_customer_ids))

#                 ai_credits_unique_result = await db.execute(ai_credits_unique_users_stmt)
#                 ai_credits_unique_users_count = ai_credits_unique_result.scalar() or 0

#                 # Compute AI Credits Total Purchases exactly like compute_gmv_totals
#                 ai_total_stmt = (
#                     select(func.count(Payment.id))
#                     .select_from(Payment)
#                     .outerjoin(Client, Payment.customer_id == Client.client_id)
#                     .where(Payment.status == "captured")
#                     .where(or_(
#                         func.json_extract(Payment.payment_metadata, '$.flow') == 'food_scanner_credits',
#                         func.json_extract(Payment.payment_metadata, '$.flow') == 'food_scanner_credits_razorpay'
#                     ))
#                     .where(~Client.contact.in_(EXCLUDED_CONTACTS))
#                 )
#                 if start_date:
#                     ai_total_stmt = ai_total_stmt.where(func.date(Payment.captured_at) >= start_date_obj)
#                 if end_date:
#                     ai_total_stmt = ai_total_stmt.where(func.date(Payment.captured_at) <= end_date_obj)
#                 if location_client_ids:
#                     location_customer_ids = {str(cid) for cid in location_client_ids}
#                     ai_total_stmt = ai_total_stmt.where(Payment.customer_id.in_(location_customer_ids))
                    
#                 ai_total_result = await db.execute(ai_total_stmt)
#                 ai_credits_total_purchases = ai_total_result.scalar() or 0

#                 # Build purchases over time (graph only)
#                 for row in ai_credits_results:
#                     date_key = row.purchase_date.isoformat() if row.purchase_date else None
#                     if date_key:
#                         if date_key not in all_purchases_over_time:
#                             all_purchases_over_time[date_key] = 0
#                         all_purchases_over_time[date_key] += row.purchase_count
#                         category_breakdown["ai_credits"]["purchases_over_time"].append({
#                             "date": date_key,
#                             "purchases": row.purchase_count
#                         })

#                 category_breakdown["ai_credits"]["purchases"] = ai_credits_total_purchases
#                 category_breakdown["ai_credits"]["unique_users"] = ai_credits_unique_users_count

#             except Exception as e:
#                 import logging
#                 logging.error(f"Purchase analytics - AI Credits error: {str(e)}")
#                 pass

#         # 3.6. AI DIET COACH PURCHASES
#         if (not source or source == "ai_diet_coach") and not gym_id:
#             try:
#                 ai_diet_coach_stmt = (
#                     select(
#                         func.date(Payment.captured_at).label('purchase_date'),
#                         func.count().label('purchase_count'),
#                         func.count(distinct(Payment.customer_id)).label('unique_users')
#                     )
#                     .select_from(Payment)
#                     .outerjoin(Client, Payment.customer_id == Client.client_id)
#                     .where(Payment.status == "captured")
#                     .where(func.json_extract(Payment.payment_metadata, '$.flow') == 'ai_diet_coach')
#                     .where(func.date(Payment.captured_at) >= start_date_obj)
#                     .where(func.date(Payment.captured_at) <= end_date_obj)
#                     .where(~Client.contact.in_(EXCLUDED_CONTACTS))
#                 )

#                 if location_client_ids:
#                     location_customer_ids = {str(cid) for cid in location_client_ids}
#                     ai_diet_coach_stmt = ai_diet_coach_stmt.where(Payment.customer_id.in_(location_customer_ids))

#                 ai_diet_coach_stmt = ai_diet_coach_stmt.group_by(func.date(Payment.captured_at))

#                 result = await db.execute(ai_diet_coach_stmt)
#                 ai_diet_coach_results = result.all()

#                 # Get total unique users across all dates
#                 ai_diet_coach_unique_users_stmt = (
#                     select(func.count(distinct(Payment.customer_id)))
#                     .select_from(Payment)
#                     .outerjoin(Client, Payment.customer_id == Client.client_id)
#                     .where(Payment.status == "captured")
#                     .where(func.json_extract(Payment.payment_metadata, '$.flow') == 'ai_diet_coach')
#                     .where(func.date(Payment.captured_at) >= start_date_obj)
#                     .where(func.date(Payment.captured_at) <= end_date_obj)
#                     .where(~Client.contact.in_(EXCLUDED_CONTACTS))
#                 )

#                 if location_client_ids:
#                     location_customer_ids = {str(cid) for cid in location_client_ids}
#                     ai_diet_coach_unique_users_stmt = ai_diet_coach_unique_users_stmt.where(Payment.customer_id.in_(location_customer_ids))

#                 ai_diet_coach_unique_result = await db.execute(ai_diet_coach_unique_users_stmt)
#                 ai_diet_coach_unique_users_count = ai_diet_coach_unique_result.scalar() or 0

#                 ai_diet_total_stmt = (
#                     select(func.count(Payment.id))
#                     .select_from(Payment)
#                     .outerjoin(Client, Payment.customer_id == Client.client_id)
#                     .where(Payment.status == "captured")
#                     .where(func.json_extract(Payment.payment_metadata, '$.flow') == 'ai_diet_coach')
#                     .where(~Client.contact.in_(EXCLUDED_CONTACTS))
#                 )
#                 if start_date:
#                     ai_diet_total_stmt = ai_diet_total_stmt.where(func.date(Payment.captured_at) >= start_date_obj)
#                 if end_date:
#                     ai_diet_total_stmt = ai_diet_total_stmt.where(func.date(Payment.captured_at) <= end_date_obj)
#                 if location_client_ids:
#                     location_customer_ids = {str(cid) for cid in location_client_ids}
#                     ai_diet_total_stmt = ai_diet_total_stmt.where(Payment.customer_id.in_(location_customer_ids))
                    
#                 ai_diet_total_result = await db.execute(ai_diet_total_stmt)
#                 ai_diet_coach_total_purchases = ai_diet_total_result.scalar() or 0

#                 for row in ai_diet_coach_results:
#                     date_key = row.purchase_date.isoformat() if row.purchase_date else None
#                     if date_key:
#                         if date_key not in all_purchases_over_time:
#                             all_purchases_over_time[date_key] = 0
#                         all_purchases_over_time[date_key] += row.purchase_count
#                         category_breakdown["ai_diet_coach"]["purchases_over_time"].append({
#                             "date": date_key,
#                             "purchases": row.purchase_count
#                         })

#                 category_breakdown["ai_diet_coach"]["purchases"] = ai_diet_coach_total_purchases
#                 category_breakdown["ai_diet_coach"]["unique_users"] = ai_diet_coach_unique_users_count

#             except Exception as e:
#                 import logging
#                 logging.error(f"Purchase analytics - AI Diet Coach error: {str(e)}")
#                 pass

#         # 4. GYM MEMBERSHIP PURCHASES
#         # GMV fix: SQL EXISTS subquery with Gym JOIN + Client JOIN (replaces Python-loop in-memory filtering)
#         if not source or source == "gym_membership":
#             try:
#                 gym_meta_cond = or_(
#                     func.json_unquote(func.json_extract(Order.order_metadata, "$.audit.source")) == "dailypass_checkout_api",
#                     func.json_unquote(func.json_extract(Order.order_metadata, "$.order_info.flow")) == "unified_gym_membership_with_sub",
#                     func.json_unquote(func.json_extract(Order.order_metadata, "$.order_info.flow")) == "unified_gym_membership_with_free_fittbot",
#                     func.json_unquote(func.json_extract(Order.order_metadata, "$.order_info.flow")) == "gym_membership_with_bonus_credits",
#                     func.json_unquote(func.json_extract(Order.order_metadata, "$.order_info.flow")) == "personal_training_with_bonus_credits"
#                 )

#                 # EXISTS on OrderItem + Gym JOIN — confirms gym physically exists
#                 gym_exists_cond = (
#                     select(1)
#                     .select_from(OrderItem)
#                     .join(Gym, Gym.gym_id == cast(OrderItem.gym_id, Integer))
#                     .where(
#                         OrderItem.order_id == Order.id,
#                         OrderItem.gym_id.isnot(None),
#                         OrderItem.gym_id != "",
#                         OrderItem.gym_id != "1"
#                     )
#                     .exists()
#                 )

#                 gm_base_conditions = [
#                     Payment.status == "captured",
#                     Order.status == "paid",
#                     Order.customer_id.isnot(None),
#                     gym_meta_cond,
#                     gym_exists_cond,
#                     func.date(Payment.captured_at) >= start_date_obj,
#                     func.date(Payment.captured_at) <= end_date_obj,
#                 ]

#                 if gym_id:
#                     gm_base_conditions.append(
#                         select(1).select_from(OrderItem)
#                         .where(OrderItem.order_id == Order.id, OrderItem.gym_id == str(gym_id))
#                         .exists()
#                     )

#                 if location_client_ids:
#                     location_customer_ids = {str(cid) for cid in location_client_ids}
#                     gm_base_conditions.append(Order.customer_id.in_(location_customer_ids))

#                 # Deduped subquery for counts
#                 gm_subq = (
#                     select(
#                         Order.id.label("order_id"),
#                         Payment.captured_at.label("captured_at"),
#                         Order.customer_id.label("customer_id")
#                     )
#                     .select_from(Payment)
#                     .join(Order, Order.id == Payment.order_id)
#                     .join(Client, Client.client_id == cast(Order.customer_id, Integer))
#                     .where(*gm_base_conditions)
#                     .distinct()
#                     .subquery()
#                 )

#                 gm_total_stmt = select(
#                     func.count(gm_subq.c.order_id).label("purchases"),
#                     func.count(distinct(gm_subq.c.customer_id)).label("unique_users")
#                 ).select_from(gm_subq)

#                 gm_total_result = await db.execute(gm_total_stmt)
#                 gm_total_row = gm_total_result.one()
#                 category_breakdown["gym_membership"]["purchases"] = gm_total_row.purchases or 0
#                 category_breakdown["gym_membership"]["unique_users"] = gm_total_row.unique_users or 0

#                 # Time-series
#                 gm_time_stmt = (
#                     select(
#                         func.date(gm_subq.c.captured_at).label("purchase_date"),
#                         func.count().label("purchase_count")
#                     )
#                     .select_from(gm_subq)
#                     .group_by(func.date(gm_subq.c.captured_at))
#                 )
#                 gm_time_result = await db.execute(gm_time_stmt)
#                 for row in gm_time_result.all():
#                     date_key = row.purchase_date.isoformat() if row.purchase_date else None
#                     if date_key:
#                         if date_key not in all_purchases_over_time:
#                             all_purchases_over_time[date_key] = 0
#                         all_purchases_over_time[date_key] += row.purchase_count
#                         category_breakdown["gym_membership"]["purchases_over_time"].append({
#                             "date": date_key,
#                             "purchases": row.purchase_count
#                         })
#                 category_breakdown["gym_membership"]["purchases_over_time"].sort(key=lambda x: x["date"])

#                 # Location tracking — re-added: get Client.location for each gym_membership order
#                 if not location or location == "all":
#                     try:
#                         gm_customer_ids_stmt = select(gm_subq.c.customer_id).select_from(gm_subq)
#                         gm_customer_ids_result = await db.execute(gm_customer_ids_stmt)
#                         gm_customer_ids = list(set([str(row[0]) for row in gm_customer_ids_result.all() if row[0]]))

#                         if gm_customer_ids:
#                             gm_loc_stmt = select(Client.client_id, Client.location).where(
#                                 Client.client_id.in_(gm_customer_ids),
#                                 Client.location.isnot(None),
#                                 Client.location != ''
#                             )
#                             gm_loc_result = await db.execute(gm_loc_stmt)
#                             gm_client_locations = {str(row[0]): row[1] for row in gm_loc_result.all()}

#                             for cid in gm_customer_ids:
#                                 if cid in gm_client_locations:
#                                     normalized_loc = gm_client_locations[cid].strip().replace(' ', '_')
#                                     if normalized_loc not in location_purchases:
#                                         location_purchases[normalized_loc] = 0
#                                     location_purchases[normalized_loc] += 1
                        
#                         # GMV fix: Also track gym-wise purchases for Gym Memberships
#                         gm_gym_stmt = (
#                             select(
#                                 OrderItem.gym_id,
#                                 func.count().label('purchase_count')
#                             )
#                             .select_from(Payment)
#                             .join(Order, Order.id == Payment.order_id)
#                             .join(OrderItem, OrderItem.order_id == Order.id)
#                             .join(Gym, Gym.gym_id == cast(OrderItem.gym_id, Integer))
#                             .where(
#                                 Payment.status == "captured",
#                                 Order.status == "paid",
#                                 OrderItem.gym_id.isnot(None),
#                                 OrderItem.gym_id != "1",
#                                 func.date(Payment.captured_at) >= start_date_obj,
#                                 func.date(Payment.captured_at) <= end_date_obj,
#                                 gym_meta_cond
#                             )
#                         )
                        
#                         if location_client_ids:
#                             gm_gym_stmt = gm_gym_stmt.where(Order.customer_id.in_(location_client_ids))
                            
#                         gm_gym_stmt = gm_gym_stmt.group_by(OrderItem.gym_id)
#                         gm_gym_result = await db.execute(gm_gym_stmt)
#                         for row in gm_gym_result.all():
#                             try:
#                                 g_id = int(row.gym_id)
#                                 if g_id not in gym_purchases:
#                                     gym_purchases[g_id] = 0
#                                 gym_purchases[g_id] += row.purchase_count
#                             except (ValueError, TypeError):
#                                 pass

#                     except Exception as e:
#                         import logging
#                         logging.error(f"Error getting location-wise gym membership purchases: {str(e)}")

#             except Exception as e:
#                 import logging
#                 logging.error(f"Purchase analytics - Gym Membership error: {str(e)}")
#                 pass

#         # Convert all purchases over time to sorted array
#         purchases_over_time = [
#             {
#                 "date": date,
#                 "purchases": count
#             }
#             for date, count in sorted(all_purchases_over_time.items())
#         ]

#         # Calculate total purchases across all categories
#         total_purchases = sum(cat_data["purchases"] for cat_data in category_breakdown.values())

#         # Debug logging
#         import logging
#         logging.info(f"Purchase analytics total_purchases: {total_purchases}, gym_purchases: {gym_purchases}")

#         # Build gym breakdown
#         gym_breakdown = []
#         if not gym_id and gym_purchases:
#             # Build full gym breakdown when no gym filter is applied
#             gym_names = {}
#             gym_ids = list(gym_purchases.keys())
#             gym_stmt = select(Gym.gym_id, Gym.name).where(Gym.gym_id.in_(gym_ids))
#             gym_result = await db.execute(gym_stmt)
#             for gym_id_val, gym_name in gym_result.all():
#                 gym_names[gym_id_val] = gym_name

#             gym_breakdown = [
#                 {
#                     "gym_id": gym_id,
#                     "gym_name": gym_names.get(gym_id, f"Gym {gym_id}"),
#                     "revenue": gym_purchases[gym_id]
#                 }
#                 for gym_id in sorted(gym_purchases.keys(), key=lambda x: gym_purchases[x], reverse=True)
#             ]
#         elif gym_id:
#             # When gym filter is applied, include the filtered gym in breakdown
#             gym_names = {}
#             gym_stmt = select(Gym.gym_id, Gym.name).where(Gym.gym_id == gym_id)
#             gym_result = await db.execute(gym_stmt)
#             for gym_id_val, gym_name in gym_result.all():
#                 gym_names[gym_id_val] = gym_name

#             # Use the gym_purchases count if available, otherwise 0
#             gym_purchases_count = gym_purchases.get(gym_id, 0)

#             gym_breakdown = [{
#                 "gym_id": gym_id,
#                 "gym_name": gym_names.get(gym_id, f"Gym {gym_id}"),
#                 "revenue": gym_purchases_count
#             }]
#             logging.info(f"Gym breakdown for filtered gym {gym_id}: {gym_breakdown}")

#         # Build location breakdown
#         location_breakdown = []
#         if location_purchases:
#             # Sort locations by purchase count (descending)
#             location_breakdown = [
#                 {
#                     "location": loc,
#                     "purchases": count
#                 }
#                 for loc, count in sorted(location_purchases.items(), key=lambda x: x[1], reverse=True)
#             ]
#             logging.info(f"Location breakdown: {location_breakdown}")

#         # Build revenue by city breakdown — uses same GMV source logic as compute_gmv_totals()
#         # Sums revenue from all 5 sources grouped by Gym.city
#         revenue_by_city = []
#         try:
#             EXCLUDED_CONTACTS_SET = ["7373675762", "9486987082", "8667458723", "9840633149", "8667427956", "8667488723", "7975847236"]
#             city_revenue_map = {}  # city -> total_revenue (rupees)

#             # ── 1. Daily Pass ────────────────────────────────────────────────────
#             try:
#                 _dp_session = get_dailypass_session()
#                 dp_city_q = (
#                     _dp_session.query(
#                         Gym.city.label("city"),
#                         func.coalesce(func.sum(Payment.amount_minor / 100.0), 0).label("revenue")
#                     )
#                     .join(Gym, Gym.gym_id == func.cast(DailyPass.gym_id, Integer), isouter=False)
#                     .outerjoin(Payment, DailyPass.payment_id == Payment.provider_payment_id)
#                     .filter(
#                         DailyPass.gym_id != "1",
#                         func.date(DailyPass.created_at) >= start_date_obj,
#                         func.date(DailyPass.created_at) <= end_date_obj,
#                     )
#                     .group_by(Gym.city)
#                 )
#                 if gym_id:
#                     dp_city_q = dp_city_q.filter(DailyPass.gym_id == str(gym_id))
#                 for row in dp_city_q.all():
#                     city = (row.city or "Unknown").strip()
#                     city_revenue_map[city] = city_revenue_map.get(city, 0) + float(row.revenue or 0)
#                 _dp_session.close()
#             except Exception as e:
#                 logging.error(f"[RevCity] daily_pass error: {e}")

#             # ── 2. Sessions (SessionPurchase) ────────────────────────────────────
#             try:
#                 sess_city_stmt = (
#                     select(
#                         func.coalesce(Gym.city, "Unknown").label("city"),
#                         func.coalesce(func.sum(SessionPurchase.payable_rupees), 0).label("revenue")
#                     )
#                     .select_from(SessionPurchase)
#                     .join(Gym, SessionPurchase.gym_id == Gym.gym_id)
#                     .where(
#                         SessionPurchase.status == "paid",
#                         SessionPurchase.gym_id != 1,
#                         func.date(SessionPurchase.created_at) >= start_date_obj,
#                         func.date(SessionPurchase.created_at) <= end_date_obj,
#                     )
#                     .group_by(Gym.city)
#                 )
#                 if gym_id:
#                     sess_city_stmt = sess_city_stmt.where(SessionPurchase.gym_id == gym_id)
#                 sess_city_result = await db.execute(sess_city_stmt)
#                 for row in sess_city_result.all():
#                     city = (row.city or "Unknown").strip()
#                     city_revenue_map[city] = city_revenue_map.get(city, 0) + float(row.revenue or 0)
#             except Exception as e:
#                 logging.error(f"[RevCity] sessions error: {e}")

#             # ── 3. Nutrition Plans ───────────────────────────────────────────────
#             # No city/gym dimension — assign to "App" bucket
#             try:
#                 nutri_city_stmt = (
#                     select(func.coalesce(func.sum(Payment.amount_minor / 100.0), 0).label("revenue"))
#                     .select_from(Payment)
#                     .outerjoin(Client, Payment.customer_id == Client.client_id)
#                     .join(
#                         NutritionEligibility,
#                         cast(Payment.order_id, String) == NutritionEligibility.source_id
#                     )
#                     .where(
#                         NutritionEligibility.source_type == "fymble_purchase",
#                         Payment.status == "captured",
#                         or_(
#                             func.json_extract(Payment.payment_metadata, "$.flow") == "nutrition_purchase_googleplay",
#                             func.json_extract(Payment.payment_metadata, "$.flow") == "nutrition_package_razorpay",
#                             func.json_extract(Payment.payment_metadata, "$.flow") == "basic_nutrition_plan",
#                             func.json_extract(Payment.payment_metadata, "$.flow") == "expert_nutrition_plan",
#                             func.json_extract(Payment.payment_metadata, "$.flow") == "elite_nutrition_plan"
#                         ),
#                         func.date(Payment.captured_at) >= start_date_obj,
#                         func.date(Payment.captured_at) <= end_date_obj,
#                         ~Client.contact.in_(EXCLUDED_CONTACTS_SET),
#                     )
#                 )
#                 nutri_result = await db.execute(nutri_city_stmt)
#                 nutri_rev = float(nutri_result.scalar() or 0)
#                 if nutri_rev > 0:
#                     city_revenue_map["App"] = city_revenue_map.get("App", 0) + nutri_rev
#             except Exception as e:
#                 logging.error(f"[RevCity] nutrition error: {e}")

#             # ── 4. AI Credits ────────────────────────────────────────────────────
#             # No city/gym dimension — assign to "App" bucket
#             try:
#                 ai_city_stmt = (
#                     select(func.coalesce(func.sum(Payment.amount_minor / 100.0), 0).label("revenue"))
#                     .select_from(Payment)
#                     .outerjoin(Client, Payment.customer_id == Client.client_id)
#                     .where(
#                         Payment.status == "captured",
#                         or_(
#                             func.json_extract(Payment.payment_metadata, "$.flow") == "food_scanner_credits",
#                             func.json_extract(Payment.payment_metadata, "$.flow") == "food_scanner_credits_razorpay"
#                         ),
#                         func.date(Payment.captured_at) >= start_date_obj,
#                         func.date(Payment.captured_at) <= end_date_obj,
#                         ~Client.contact.in_(EXCLUDED_CONTACTS_SET),
#                     )
#                 )
#                 ai_result = await db.execute(ai_city_stmt)
#                 ai_rev = float(ai_result.scalar() or 0)
#                 if ai_rev > 0:
#                     city_revenue_map["App"] = city_revenue_map.get("App", 0) + ai_rev
#             except Exception as e:
#                 logging.error(f"[RevCity] ai_credits error: {e}")

#             # ── 5. Gym Membership ────────────────────────────────────────────────
#             try:
#                 gm_meta_cond = or_(
#                     func.json_unquote(func.json_extract(Order.order_metadata, "$.audit.source")) == "dailypass_checkout_api",
#                     func.json_unquote(func.json_extract(Order.order_metadata, "$.order_info.flow")) == "unified_gym_membership_with_sub",
#                     func.json_unquote(func.json_extract(Order.order_metadata, "$.order_info.flow")) == "unified_gym_membership_with_free_fittbot",
#                 )
#                 gm_exists = (
#                     select(1)
#                     .select_from(OrderItem)
#                     .join(Gym, Gym.gym_id == cast(OrderItem.gym_id, Integer))
#                     .where(
#                         OrderItem.order_id == Order.id,
#                         OrderItem.gym_id.isnot(None),
#                         OrderItem.gym_id != "",
#                         OrderItem.gym_id != "1",
#                     )
#                     .exists()
#                 )
#                 gm_city_conditions = [
#                     Payment.status == "captured",
#                     Order.status == "paid",
#                     Order.customer_id.isnot(None),
#                     gm_meta_cond,
#                     gm_exists,
#                     func.date(Payment.captured_at) >= start_date_obj,
#                     func.date(Payment.captured_at) <= end_date_obj,
#                 ]
#                 if gym_id:
#                     gm_city_conditions.append(
#                         select(1).select_from(OrderItem)
#                         .where(OrderItem.order_id == Order.id, OrderItem.gym_id == str(gym_id))
#                         .exists()
#                     )

#                 # Join through OrderItem to get the gym, then Gym.city
#                 gm_city_subq = (
#                     select(
#                         Order.id.label("order_id"),
#                         Order.gross_amount_minor.label("gross_amount_minor"),
#                         OrderItem.gym_id.label("item_gym_id"),
#                     )
#                     .select_from(Payment)
#                     .join(Order, Order.id == Payment.order_id)
#                     .join(Client, Client.client_id == cast(Order.customer_id, Integer))
#                     .join(OrderItem, and_(
#                         OrderItem.order_id == Order.id,
#                         OrderItem.gym_id.isnot(None),
#                         OrderItem.gym_id != "",
#                         OrderItem.gym_id != "1",
#                     ))
#                     .where(*gm_city_conditions)
#                     .distinct()
#                     .subquery()
#                 )

#                 gm_city_stmt = (
#                     select(
#                         func.coalesce(Gym.city, "Unknown").label("city"),
#                         func.coalesce(func.sum(gm_city_subq.c.gross_amount_minor / 100.0), 0).label("revenue"),
#                     )
#                     .select_from(gm_city_subq)
#                     .join(Gym, Gym.gym_id == cast(gm_city_subq.c.item_gym_id, Integer))
#                     .group_by(Gym.city)
#                 )
#                 gm_city_result = await db.execute(gm_city_stmt)
#                 for row in gm_city_result.all():
#                     city = (row.city or "Unknown").strip()
#                     city_revenue_map[city] = city_revenue_map.get(city, 0) + float(row.revenue or 0)
#             except Exception as e:
#                 logging.error(f"[RevCity] gym_membership error: {e}")

#             # Build sorted output
#             revenue_by_city = sorted(
#                 [{"city": city, "amount": round(amt, 2)} for city, amt in city_revenue_map.items() if city],
#                 key=lambda x: x["amount"],
#                 reverse=True
#             )[:20]

#             logging.info(f"Final revenue_by_city: {revenue_by_city}")

#         except Exception as e:
#             logging.error(f"Error building revenue_by_city: {str(e)}")
#             import traceback
#             traceback.print_exc()


#         # Build separate stable gym list for filter dropdown (gyms with purchases in current timeframe)
#         all_available_gyms = []
#         if gym_purchases:
#             all_gym_ids = list(gym_purchases.keys())
#             all_gym_stmt = select(Gym.gym_id, Gym.name).where(Gym.gym_id.in_(all_gym_ids))
#             all_gym_res = await db.execute(all_gym_stmt)
#             gym_names_map = {row.gym_id: row.name for row in all_gym_res.all()}
            
#             all_available_gyms = [
#                 {
#                     "gym_id": gid,
#                     "gym_name": gym_names_map.get(gid, f"Gym {gid}")
#                 }
#                 for gid in sorted(gym_purchases.keys(), key=lambda x: gym_purchases[x], reverse=True)
#             ]

#         analytics_data = {
#             "totalPurchases": total_purchases,
#             "categoryBreakdown": category_breakdown,
#             "purchasesOverTime": purchases_over_time,
#             "gymBreakdown": gym_breakdown,
#             "availableGyms": all_available_gyms,  # Providing stable list for filters
#             "locationBreakdown": location_breakdown,
#             "revenueByCity": revenue_by_city,
#             "filters": {
#                 "startDate": start_date if start_date else "All Time",
#                 "endDate": end_date if end_date else "All Time",
#                 "source": source or "all",
#                 "gymId": gym_id or "all",
#                 "location": location or "all"
#             }
#         }

#         return {
#             "success": True,
#             "data": analytics_data,
#             "message": "Purchase analytics fetched successfully"
#         }

#     except ValueError as e:
#         raise HTTPException(status_code=400, detail=f"Invalid date format: {str(e)}")
#     except Exception as e:
#         raise HTTPException(status_code=500, detail=str(e))

